import json, csv, io
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.utils import timezone
from django.shortcuts import get_object_or_404, render, redirect
from django.views.decorators.http import require_POST
from inventory.models import Issue, Item, Room, Category, RoomBooking, Purchase, Vendor, Department, RoomBookingCredentials, RoomBookingRequest, RoomCancellationRequest, Brand
from core.models import UserProfile
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
from decimal import Decimal, InvalidOperation
from django.views.generic import FormView
from django.views import View
from django.contrib import messages
from django.urls import reverse
from inventory.forms.room_incharge import ExcelUploadForm
from django.db import models
from inventory.models import SystemComponent as SC
import re as _re

def _extract_docx_structured(raw_bytes):
    """
    Parse a .docx binary and return:
      {
        'blocks': [
            {'type': 'paragraph', 'text': '...'},
            {'type': 'table', 'rows': [['cell','cell',...], ...]},
            ...
        ],
        'plain_text': 'flattened string for caching / PDF fallback'
      }

    Covers:
      1. Body paragraphs  (headings, normal text, list items)
      2. Tables           (preserves row/column structure)
      3. Text boxes       (<w:txbxContent> drawing shapes)

    The document's XML body is walked in document order so blocks come out
    in the same sequence they appear on the page.
    """
    import io as _io
    from docx import Document as _Doc
    from docx.oxml.ns import qn as _qn

    doc = _Doc(_io.BytesIO(raw_bytes))
    blocks = []
    plain_lines = []

    # Walk the body XML in document order so paragraphs and tables are
    # interleaved correctly (doc.paragraphs and doc.tables are separate lists
    # and lose positional context relative to each other).
    body = doc.element.body
    for child in body:
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag

        if tag == 'p':
            # Plain paragraph (heading, body text, list item, etc.)
            text = ''.join(
                node.text or ''
                for node in child.iter(_qn('w:t'))
            ).strip()
            if text:
                blocks.append({'type': 'paragraph', 'text': text})
                plain_lines.append(text)

        elif tag == 'tbl':
            # Table — collect all rows and cells
            rows_data = []
            for tr in child.iter(_qn('w:tr')):
                row_cells = []
                for tc in tr.iter(_qn('w:tc')):
                    # Each cell may have multiple paragraphs
                    cell_text = ' '.join(
                        ''.join(node.text or '' for node in p.iter(_qn('w:t'))).strip()
                        for p in tc.iter(_qn('w:p'))
                    ).strip()
                    row_cells.append(cell_text)
                if any(c for c in row_cells):
                    rows_data.append(row_cells)

            if rows_data:
                blocks.append({'type': 'table', 'rows': rows_data})
                # Flatten table for plain_text cache
                for row in rows_data:
                    plain_lines.append(' | '.join(row))

        elif tag == 'txbxContent':
            # Text box
            text = ''.join(
                node.text or ''
                for node in child.iter(_qn('w:t'))
            ).strip()
            if text:
                blocks.append({'type': 'paragraph', 'text': text})
                plain_lines.append(text)

    # Also catch text boxes nested anywhere deeper in the body
    ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
    for txbx in body.findall('.//w:txbxContent', ns):
        # Only pick up ones NOT already in the top-level walk above
        if txbx.getparent() is not None and txbx.getparent().getparent() is not body:
            text = ''.join(
                node.text or ''
                for node in txbx.iter(_qn('w:t'))
            ).strip()
            if text:
                blocks.append({'type': 'paragraph', 'text': text})
                plain_lines.append(text)

    plain_text = '\n'.join(plain_lines) if plain_lines else None
    return {'blocks': blocks, 'plain_text': plain_text}

class CentralAdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        profile = getattr(self.request.user, 'profile', None)
        return profile and profile.is_central_admin or profile.is_sub_admin

class AuraDashboardView(LoginRequiredMixin, CentralAdminRequiredMixin, TemplateView):
    template_name = 'central_admin/aura_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.profile.org

        context['total_issues']    = Issue.objects.all().count()
        context['escalated_count'] = Issue.objects.filter(status='escalated').count()
        context['total_items']     = Item.objects.all().count()
        context['booking_credentials'] = RoomBookingCredentials.objects.all().order_by('email')

        # Confirmed (approved) bookings
        try:
            context['active_bookings'] = RoomBooking.objects.filter(
                room__organisation=org,
                end_datetime__gte=timezone.now()
            ).count()
        except Exception:
            context['active_bookings'] = 0

        # Pending approval counts for dashboard alert banner
        try:
            from inventory.models import RoomBookingRequest, RoomCancellationRequest
            context['pending_booking_requests'] = RoomBookingRequest.objects.filter(status='pending').count()
            context['pending_cancel_requests']  = RoomCancellationRequest.objects.filter(status='pending').count()
        except Exception:
            context['pending_booking_requests'] = 0
            context['pending_cancel_requests']  = 0

        return context

def aura_analytics_data(request):
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    org = request.user.profile.org
    
    # Live data for Charts
    issue_stats = Issue.objects.all().values('status').annotate(count=Count('id'))
    category_stats = Item.objects.all().values('category__category_name').annotate(count=Count('id'))
    
    total_rooms = Room.objects.all().count()
    booked_today = 0
    
    # Defensive query for Room Utilization
    try:
        booked_today = RoomBooking.objects.filter(
            # room__organisation=org,
            start_datetime__date=timezone.now().date()
        ).values('room').distinct().count()
    except Exception:
        pass 

    return JsonResponse({
        'issue_labels': [s['status'].replace('_', ' ').title() for s in issue_stats],
        'issue_series': [s['count'] for s in issue_stats],
        'cat_labels': [s['category__category_name'] for s in category_stats],
        'cat_series': [s['count'] for s in category_stats],
        'room_util': [booked_today, max(0, total_rooms - booked_today)]
    })

def aura_data_manager(request):
    """
    Fetches AURA report data for all modules with module and date filtering.
    Accessible by both central admin and sub-admin (sub-admin gets view-only).
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    model_name = request.GET.get('model')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    model_map = {
        'issues': Issue, 
        'items': Item, 
        'rooms': Room, 
        'bookings': RoomBooking, 
        'purchases': Purchase,
        'vendors': Vendor, 
        'departments': Department,
        'credentials': RoomBookingCredentials,
    }
    
    model = model_map.get(model_name)
    if not model:
        return JsonResponse({'error': 'Invalid model'}, status=400)
    
    qs = model.objects.all().order_by('-id')

    # Apply Date Filtering across all applicable modules
    if date_from and date_to:
        if model_name == 'bookings':
            qs = qs.filter(start_datetime__date__range=[date_from, date_to])
        elif model_name in ['issues', 'items', 'rooms', 'purchases']:
            qs = qs.filter(created_on__date__range=[date_from, date_to])

    data = []
    for obj in qs:
        row = {'id': obj.id}
        try:
            if model_name == 'rooms':
                row['label_head'] = "Room"
                row['label'] = f"{obj.room_name}<br><small>Incharge: {obj.incharge}</small>"
                row['detail_head'] = "Metadata"
                row['detail'] = f"Category: {obj.get_room_category_display()} | Capacity: {obj.capacity}"
            elif model_name == 'bookings':
                # Report Generator Fields (Multi-line)
                row['label_head'] = "Room Booked"
                row['label'] = f"{obj.faculty_name}<br><small>{obj.faculty_email}</small>"
                row['detail_head'] = "Metadata"
                start_local = timezone.localtime(obj.start_datetime)
                end_local = timezone.localtime(obj.end_datetime)
                row['detail'] = f"Room: {obj.room.room_name}<br>Date: {start_local.strftime('%d %b, %Y')}<br>Time: {start_local.strftime('%H:%M')} - {end_local.strftime('%H:%M')}"
                
                # SPECIFIC KEYS FOR ROOM BOOKING MANAGER MODAL
                row['room_name'] = obj.room.room_name
                row['faculty_name'] = obj.faculty_name
                row['faculty_email'] = obj.faculty_email
                row['schedule'] = f"{start_local.strftime('%d %b, %Y')} | {start_local.strftime('%H:%M')} – {end_local.strftime('%H:%M')}"
            elif model_name == 'issues':
                row['label_head'] = "Issue"
                row['label'] = f"{obj.subject}"
                row['detail_head'] = "Metadata"
                assigned = obj.assigned_to.user.get_full_name() if obj.assigned_to else "N/A"
                row['detail'] = f"Email: {obj.reporter_email}<br>Ticket ID: {obj.ticket_id}<br>Status: {obj.status}<br>Room: {obj.room.room_name}<br>Assigned: {assigned}"
            elif model_name == 'items':
                from inventory.models import AssetTag
                row['label_head'] = "Items"
                row['label'] = f"{obj.item_name}"
                row['detail_head'] = "Metadata"
                room_name = obj.room.room_name if obj.room else 'Master Inventory'
                prod_code = obj.product_code or '—'
                # Asset tag range for this room
                tag_range = '—'
                if obj.room and obj.product_code:
                    tags = AssetTag.objects.filter(
                        item_name=obj.item_name,
                        assigned_room=obj.room
                    ).order_by('tag_id')
                    if tags.exists():
                        first_tag = tags.first().tag_id
                        last_tag = tags.last().tag_id
                        tag_range = f"{first_tag} → {last_tag}"
                row['detail'] = f"Room: {room_name}<br>Product Code: {prod_code}<br>Asset Tags: {tag_range}<br>Qty: {obj.total_count}<br>Available: {obj.available_count}<br>In Use: {obj.in_use}"
            elif model_name == 'purchases':
                row['label_head'] = "Purchase ID"
                room_name = obj.room.room_name if obj.room else 'No Room Assigned'
                row['label'] = f"{obj.purchase_id or 'Pending ID'}<br>Room: {room_name}"
                row['detail_head'] = "Metadata"
                item_name = obj.item.item_name if obj.item else 'N/A'
                vendor_name = obj.vendor.vendor_name if obj.vendor else 'No Vendor'
                row['detail'] = f"Item: {item_name}<br>Vendor: {vendor_name}<br>Status: {obj.status.title()}"
            elif model_name == 'vendors':
                row['label_head'] = "Vendors"
                row['label'] = f"{obj.vendor_name}"
                row['detail_head'] = "Metadata"
                row['detail'] = f"Email: {obj.email}<br>Contact: {obj.contact_number}"
            elif model_name == 'departments':
                row['label_head'] = "Department/Cell/Office"
                row['label'] = f"{obj.department_name}"
                row['detail_head'] = "Metadata"
                room_count = Room.objects.filter(department=obj).count()
                row['detail'] = f"Total Rooms: {room_count}"
            elif model_name == 'credentials':
                row['label_head'] = "Email"
                row['label'] = obj.email
                row['detail_head'] = "Designation"
                row['detail'] = obj.designation or 'Faculty'
                row['email'] = obj.email
                row['designation'] = obj.designation or 'Faculty'
                row['password'] = obj.password
            else:
                row['label_head'] = "Primary Record"
                row['label'] = str(obj)
                row['detail_head'] = "Metadata / Details"
                row['detail'] = "General Record"
        except Exception:
            row['detail'] = "N/A (Data Mismatch)"
        data.append(row)
        
    return JsonResponse({'results': data})

def aura_delete_record(request):
    if request.method == 'POST' and request.user.profile.is_central_admin:
        data = json.loads(request.body)
        model_name = data.get('model')
        record_id = data.get('id')
        
        model_map = {'issues': Issue, 'items': Item, 'rooms': Room, 'bookings': RoomBooking}
        model = model_map.get(model_name)
        obj = get_object_or_404(model, id=record_id)
        
        # Ownership check
        is_owner = False
        if model_name == 'rooms' and obj.organisation == request.user.profile.org:
            is_owner = True
        elif hasattr(obj, 'room') and obj.room.organisation == request.user.profile.org:
            is_owner = True
        elif hasattr(obj, 'organisation') and obj.organisation == request.user.profile.org:
            is_owner = True
            
        if is_owner:
            obj.delete()
            return JsonResponse({'status': 'success'})
            
    return JsonResponse({'status': 'error'}, status=403)


def get_rooms_by_category(request):
    """Get rooms filtered by category"""
    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    category = request.GET.get('category')
    org = profile.org
    
    rooms = Room.objects.filter(
        organisation=org,
        room_category=category
    ).values('id', 'room_name')
    
    return JsonResponse({
        'rooms': [{'id': r['id'], 'name': r['room_name']} for r in rooms]
    })



def get_assignment_details(request):
    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    item_id = request.GET.get('item_id')
    room_id = request.GET.get('room_id')
    org = profile.org

    master_item = get_object_or_404(Item, id=item_id, organisation=org, room__isnull=True)
    room = get_object_or_404(Room, id=room_id, organisation=org)

    # Match by item_name only — category object differs between master and room
    assigned_item = Item.objects.filter(
        organisation=org,
        room=room,
        item_name=master_item.item_name,
    ).first()

    if assigned_item:
        return JsonResponse({
            'assigned_quantity': assigned_item.total_count,
            'room_name': room.room_name,
            'item_name': master_item.item_name,
            'room_item_id': assigned_item.id,
        })
    else:
        return JsonResponse({
            'assigned_quantity': 0,
            'room_name': room.room_name,
            'item_name': master_item.item_name,
            'room_item_id': None,
        })


def get_room_inventory(request):
    """
    Returns all inventory items assigned to a specific room.
    Used by the 'View Room Inventory' modal in assign_inventory.html.
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    room_id = request.GET.get('room_id')
    if not room_id:
        return JsonResponse({'error': 'room_id is required'}, status=400)

    try:
        room = Room.objects.get(id=room_id, organisation=profile.org)
    except Room.DoesNotExist:
        return JsonResponse({'error': 'Room not found'}, status=404)

    items_qs = Item.objects.filter(
        organisation=profile.org,
        room=room,
    ).select_related('category', 'brand').order_by('item_name')

    items_data = []
    for item in items_qs:
        items_data.append({
            'item_name': item.item_name,
            'category': item.category.category_name if item.category else '—',
            'brand': item.brand.brand_name if item.brand else '—',
            'total_qty': item.total_count or 0,
            'available': item.available_count or 0,
            'in_use': item.in_use or 0,
        })

    return JsonResponse({
        'room_name': room.room_name,
        'items': items_data,
        'total_items': len(items_data),
    })


def aura_generate_report_pdf(request):
    """
    Solid fix for large datasets: Server-side PDF generation.
    Uses ReportLab and .iterator() to prevent blank pages and memory crashes.
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return HttpResponse("Unauthorized", status=403)

    # 1. Capture Filters
    model_name = request.GET.get('model')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    model_map = {
        'issues': Issue, 'items': Item, 'rooms': Room, 
        'bookings': RoomBooking, 'purchases': Purchase,
        'vendors': Vendor, 'departments': Department
    }
    
    model = model_map.get(model_name)
    if not model:
        return HttpResponse("Invalid Model", status=400)
    
    # 2. Query Data with Iterator
    qs = model.objects.all().order_by('id')
    if date_from and date_to:
        if model_name == 'bookings':
            qs = qs.filter(start_datetime__date__range=[date_from, date_to])
        elif model_name in ['issues', 'items', 'rooms', 'purchases']:
            qs = qs.filter(created_on__date__range=[date_from, date_to])
    

    # 3. Setup PDF Buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"Blixtro - Aura Generated Report: {model_name.upper()}", styles['Title']))
    elements.append(Spacer(1, 12))

    # Dynamic Headers
    headers = ["ID", "Record", "Metadata"]
    if model_name == 'rooms': headers = ["ID", "Room", "Metadata"]
    elif model_name == 'bookings': headers = ["ID", "Room Booked", "Metadata"]
    elif model_name == 'issues': headers = ["ID", "Issue", "Metadata"]
    elif model_name == 'items': headers = ["ID", "Items", "Metadata"]
    elif model_name == 'purchases': headers = ["ID", "Purchase ID", "Metadata"]
    elif model_name == 'vendors': headers = ["ID", "Vendors", "Metadata"]
    elif model_name == 'departments': headers = ["ID", "Department/Cell/Office", "Metadata"]
    
    report_data = [headers]
    
    # Efficiently loop through large data
    for obj in qs.iterator():
        # Replicate label/detail logic from aura_data_manager
        label = str(obj)
        detail = "General Record"
        
        try:
            if model_name == 'rooms':
                label = f"{obj.room_name}<br/>Incharge: {obj.incharge}"
                detail = f"Category: {obj.get_room_category_display()}<br/>Capacity: {obj.capacity}"
            elif model_name == 'bookings':
                label = f"{obj.faculty_name}<br/>{obj.faculty_email}"
                start_local = timezone.localtime(obj.start_datetime)
                end_local = timezone.localtime(obj.end_datetime)
                detail = f"Room: {obj.room.room_name}<br/>Date: {start_local.strftime('%d %b, %Y')}<br/>Time: {start_local.strftime('%H:%M')} - {end_local.strftime('%H:%M')}"
            elif model_name == 'issues':
                label = f"{obj.subject}"
                assigned = obj.assigned_to.user.get_full_name() if obj.assigned_to else "N/A"
                detail = f"Email: {obj.reporter_email}<br/>Ticket ID: {obj.ticket_id}<br/>Status: {obj.status}<br/>Room: {obj.room.room_name}<br/>Assigned: {assigned}"
            elif model_name == 'items':
                from inventory.models import AssetTag
                label = f"{obj.item_name}"
                room_name = obj.room.room_name if obj.room else 'Master Inventory'
                prod_code = obj.product_code or '—'
                tag_range = '—'
                if obj.room and obj.product_code:
                    tags = AssetTag.objects.filter(
                        item_name=obj.item_name,
                        assigned_room=obj.room
                    ).order_by('tag_id')
                    if tags.exists():
                        tag_range = f"{tags.first().tag_id} → {tags.last().tag_id}"
                detail = f"Room: {room_name}<br/>Product Code: {prod_code}<br/>Asset Tags: {tag_range}<br/>Qty: {obj.total_count}<br/>Available: {obj.available_count}<br/>In Use: {obj.in_use}"
            elif model_name == 'purchases':
                label = f"{obj.purchase_id or 'Pending ID'}<br/>Room: {obj.room.room_name if obj.room else 'No Room'}"
                item_name = obj.item.item_name if obj.item else 'N/A'
                vendor_name = obj.vendor.vendor_name if obj.vendor else 'No Vendor'
                detail = f"Item: {item_name}<br/>Vendor: {vendor_name}<br/>Status: {obj.status.title()}"
            elif model_name == 'vendors':
                label = f"{obj.vendor_name}"
                detail = f"Email: {obj.email}<br/>Contact: {obj.contact_number}"
            elif model_name == 'departments':
                label = f"{obj.department_name}"
                room_count = Room.objects.filter(department=obj).count()
                detail = f"Total Rooms: {room_count}"
        except Exception:
            detail = "Data Mismatch"

        report_data.append([
            str(obj.id),
            Paragraph(label, styles['Normal']),
            Paragraph(detail, styles['Normal'])
        ])

    # 4. Build Table
    table = Table(report_data, repeatRows=1, colWidths=[60, 250, 400])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Blixtro_AURA_{model_name}_Report.pdf"'
    return response

@require_POST
def aura_bulk_delete(request):
    """
    Bulk delete records for a specific model based on a list of IDs.
    """
    if not request.user.profile.is_central_admin:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        model_name = data.get('model')
        record_ids = data.get('ids', []) # Expects a list of IDs
        
        model_map = {
            'issues': Issue, 
            'items': Item, 
            'rooms': Room, 
            'bookings': RoomBooking,
            'purchases': Purchase,
            'credentials': RoomBookingCredentials,
        }
        
        model = model_map.get(model_name)
        if not model or not record_ids:
            return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
            
        # Filter records belonging to the user's organization and delete
        if model_name == 'credentials':
            qs = model.objects.filter(id__in=record_ids)
        elif model_name == 'rooms':
            qs = model.objects.filter(id__in=record_ids, organisation=request.user.profile.org)
        elif hasattr(model, 'room'):
            qs = model.objects.filter(id__in=record_ids, room__organisation=request.user.profile.org)
        else:
            qs = model.objects.filter(id__in=record_ids, organisation=request.user.profile.org)
            
        count = qs.count()
        qs.delete()
        
        return JsonResponse({'status': 'success', 'deleted_count': count})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ─────────────────────────────────────────────
# CREDENTIAL-SPECIFIC ENDPOINTS
# ─────────────────────────────────────────────

def credential_delete(request, pk):
    """Delete a single credential. Central admin only."""
    profile = getattr(request.user, 'profile', None)
    if not profile or not profile.is_central_admin:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
    cred = get_object_or_404(RoomBookingCredentials, pk=pk)
    cred.delete()
    return JsonResponse({'status': 'success'})


def credential_update(request, pk):
    """
    Edit a single credential (email, designation, password).
    Central admin only. Accepts POST with JSON body.
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or not profile.is_central_admin:
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        cred = get_object_or_404(RoomBookingCredentials, pk=pk)

        new_email = data.get('email', '').strip().lower()
        new_designation = data.get('designation', '').strip()
        new_password = data.get('password', '').strip()

        if not new_email:
            return JsonResponse({'status': 'error', 'message': 'Email is required'}, status=400)

        # Check uniqueness (exclude self)
        if RoomBookingCredentials.objects.filter(email=new_email).exclude(pk=pk).exists():
            return JsonResponse({'status': 'error', 'message': 'This email already exists'}, status=400)

        cred.email = new_email
        if new_designation:
            cred.designation = new_designation
        if new_password:
            cred.password = new_password
        cred.save()

        return JsonResponse({'status': 'success', 'email': cred.email, 'designation': cred.designation})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ─────────────────────────────────────────────
# TRACK BOOKING/CANCELLATION STATUS (Faculty-facing)
# ─────────────────────────────────────────────

def get_booking_status(request):
    """
    Faculty enters email + password to see the status of all their
    booking requests AND cancellation requests.
    Validates against RoomBookingCredentials (Faculty Manager).
    """
    email    = request.GET.get('email', '').strip().lower()
    password = request.GET.get('password', '').strip()

    if not email or not password:
        return JsonResponse({'error': 'Email and password are required.'}, status=400)

    # Validate credentials against Faculty Manager
    try:
        cred = RoomBookingCredentials.objects.get(email=email)
        if cred.password != password:
            return JsonResponse({'error': 'Incorrect password.'}, status=403)
    except RoomBookingCredentials.DoesNotExist:
        return JsonResponse({'error': 'This email is not registered in the Faculty Manager.'}, status=403)

    results = []

    # Booking requests for this faculty
    booking_reqs = RoomBookingRequest.objects.filter(
        faculty_email=email
    ).select_related('room').order_by('-created_on')

    for req in booking_reqs:
        results.append({
            'type':        'Booking Request',
            'room':        req.room.room_name,
            'from':        req.start_datetime.strftime('%d %b %Y, %H:%M'),
            'to':          req.end_datetime.strftime('%H:%M'),
            'purpose':     req.purpose or '—',
            'status':      req.status,      # pending / approved / rejected
            'review_note': req.review_note or '',
            'submitted':   req.created_on.strftime('%d %b %Y'),
        })

    # Cancellation requests for this faculty's confirmed bookings
    cancel_reqs = RoomCancellationRequest.objects.filter(
        faculty_email=email
    ).select_related('booking', 'booking__room').order_by('-created_on')

    for req in cancel_reqs:
        room_name = req.booking.room.room_name if req.booking else '—'
        results.append({
            'type':        'Cancellation Request',
            'room':        room_name,
            'from':        req.booking.start_datetime.strftime('%d %b %Y, %H:%M') if req.booking else '—',
            'to':          req.booking.end_datetime.strftime('%H:%M') if req.booking else '—',
            'purpose':     req.reason,
            'status':      req.status,
            'review_note': '',
            'submitted':   req.created_on.strftime('%d %b %Y'),
        })

    return JsonResponse({'requests': results})


# ─────────────────────────────────────────────
# CONFIRMED BOOKING FILES (AURA Command card)
# ─────────────────────────────────────────────

def confirmed_booking_files(request):
    """
    Returns ALL confirmed RoomBookings, showing purpose and download link if available.
    Used by the AURA "Confirmed Booking Files" command card.
    Central admin AND sub-admin can access this.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        bookings = RoomBooking.objects.select_related('room').order_by('-created_on')

        results = []
        for b in bookings:
            # has_doc is True whenever a file path is recorded — regardless of
            # whether text has been extracted yet. The actual reading/extraction
            # is deferred to get_booking_doc_text (called lazily on "View Doc"
            # click), so a silent storage failure here never hides the buttons.
            has_doc  = bool(b.requirements_doc and b.requirements_doc.name)
            doc_name = b.requirements_doc.name.split('/')[-1] if has_doc else None

            # Safe datetime localisation
            try:
                from_str = timezone.localtime(b.start_datetime).strftime('%d %b %Y, %H:%M')
                to_str   = timezone.localtime(b.end_datetime).strftime('%d %b %Y, %H:%M')
            except Exception:
                from_str = str(b.start_datetime)
                to_str   = str(b.end_datetime)

            # Inline text requirements (typed by faculty instead of uploading a doc)
            inline_text     = (b.requirements_text or '').strip()
            has_inline_text = bool(inline_text)

            results.append({
                'id':              b.id,
                'room':            b.room.room_name if b.room else '—',
                'faculty':         b.faculty_name  or '—',
                'email':           b.faculty_email or '—',
                'from':            from_str,
                'to':              to_str,
                'purpose':         b.purpose or '',
                'doc_name':        doc_name,
                # True whenever a doc file is attached
                'has_doc_text':    has_doc,
                # Inline text requirements (no file upload — typed directly)
                'has_inline_text': has_inline_text,
                'inline_text':     inline_text if has_inline_text else '',
            })

        return JsonResponse({'results': results})

    except Exception as e:
        return JsonResponse({'error': f'Failed to load booking files: {str(e)}'}, status=500)


# ─────────────────────────────────────────────
# DELETE CONFIRMED BOOKING (central admin only)
# ─────────────────────────────────────────────

@require_POST
def delete_confirmed_booking(request, booking_id):
    """
    Delete a single confirmed RoomBooking.
    Central admin only — sub-admins cannot delete bookings.
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or not profile.is_central_admin:
        return JsonResponse({'error': 'Only central admins can delete bookings.'}, status=403)

    booking = get_object_or_404(RoomBooking, id=booking_id)
    room_name = booking.room.room_name if booking.room else str(booking_id)
    try:
        booking.delete()
        return JsonResponse({'status': 'success', 'message': f'Booking for {room_name} deleted.'})
    except Exception as e:
        return JsonResponse({'error': f'Could not delete booking: {str(e)}'}, status=500)


@require_POST
def bulk_delete_confirmed_bookings(request):
    """
    Bulk-delete confirmed RoomBookings by a list of IDs.
    Central admin only — sub-admins cannot delete bookings.
    Expects JSON body: { "ids": [1, 2, 3, ...] }
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or not profile.is_central_admin:
        return JsonResponse({'error': 'Only central admins can delete bookings.'}, status=403)

    try:
        body = json.loads(request.body)
        ids  = body.get('ids', [])
        if not ids or not isinstance(ids, list):
            return JsonResponse({'error': 'No booking IDs provided.'}, status=400)

        # Only valid integers
        ids = [int(i) for i in ids if str(i).isdigit()]
        if not ids:
            return JsonResponse({'error': 'No valid booking IDs provided.'}, status=400)

        qs      = RoomBooking.objects.filter(id__in=ids)
        count   = qs.count()
        qs.delete()
        return JsonResponse({'status': 'success', 'deleted_count': count,
                             'message': f'{count} booking(s) deleted successfully.'})
    except Exception as e:
        return JsonResponse({'error': f'Bulk delete failed: {str(e)}'}, status=500)



# ─────────────────────────────────────────────
# PROXY DOWNLOAD VIEW FOR BOOKING DOCUMENTS
# ─────────────────────────────────────────────

def download_booking_doc(request, booking_id):
    """
    Proxy view that streams a booking's requirements document.

    Works for BOTH local development (local filesystem) and production
    (DigitalOcean Spaces / S3-compatible storage).

    On S3/Spaces the file object returned by the storage backend is not a
    regular file — calling .read() on it opens a network stream.  We read
    it in chunks to handle large files without exhausting memory, and we
    explicitly close the storage handle afterwards.

    Central admin and sub-admin only.
    """
    import mimetypes

    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return HttpResponse("Unauthorized", status=403)

    booking = get_object_or_404(RoomBooking, id=booking_id)

    if not booking.requirements_doc or not booking.requirements_doc.name:
        return HttpResponse("No document attached to this booking.", status=404)

    try:
        doc_field = booking.requirements_doc
        file_name = doc_field.name.split('/')[-1]

        content_type, _ = mimetypes.guess_type(file_name)
        if not content_type:
            content_type = 'application/octet-stream'

        # ── Strategy 1: try reading via the storage backend (works for both
        #   local files and private S3/Spaces objects accessed with credentials)
        try:
            storage = doc_field.storage
            with storage.open(doc_field.name, 'rb') as f:
                file_data = f.read()

            response = HttpResponse(file_data, content_type=content_type)
            response['Content-Disposition'] = (
                f'attachment; filename="{file_name}"'
            )
            return response

        except Exception as storage_err:
            # ── Strategy 2: if the storage backend can generate a URL,
            #   redirect the browser to a pre-signed / public URL so the
            #   download happens directly from the CDN/bucket.
            #   This is the fallback for very large files or unusual configs.
            try:
                url = doc_field.url  # may be a pre-signed URL on S3
                if url:
                    from django.http import HttpResponseRedirect
                    return HttpResponseRedirect(url)
            except Exception:
                pass

            # Both strategies failed — surface the original error
            raise storage_err

    except Exception as e:
        return HttpResponse(f"Failed to retrieve document: {str(e)}", status=500)


def get_booking_doc_text(request, booking_id):
    """
    Returns structured document content as JSON for the "View Doc" panel.

    Response format:
      {
        'doc_name': 'filename.docx',
        'blocks': [
            {'type': 'paragraph', 'text': '...'},
            {'type': 'table', 'rows': [['cell', 'cell'], ...]},
            ...
        ],
        'text': 'plain text cache string'   # also returned for compatibility
      }

    Covers body paragraphs, tables, and text boxes in document order.
    Skips the fast-path cache if it contains a stale sentinel value so that
    old bookings automatically get proper re-extraction.
    """
    import traceback as _tb

    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    booking = get_object_or_404(RoomBooking, id=booking_id)
    doc_name = (
        booking.requirements_doc.name.split('/')[-1]
        if booking.requirements_doc and booking.requirements_doc.name
        else 'Document'
    )

    # ── Always re-parse from the file so tables/textboxes are returned as ──────
    # structured blocks, not flat pipe-separated text rebuilt from cache.
    # The file read is fast (typically <1 MB) and ensures the View Doc panel
    # always shows real table structure.
    if not (booking.requirements_doc and booking.requirements_doc.name):
        return JsonResponse({'error': 'No document is attached to this booking.'}, status=404)

    try:
        from docx import Document  # noqa — just checking install
    except ImportError:
        return JsonResponse(
            {'error': 'python-docx is not installed. Run: pip install python-docx'},
            status=500
        )

    try:
        _storage = booking.requirements_doc.storage
        with _storage.open(booking.requirements_doc.name, 'rb') as _f:
            raw = _f.read()
    except Exception as e:
        print(f"[get_booking_doc_text] Storage read failed booking={booking_id}: {e}\n{_tb.format_exc()}")
        # Fall back to cached plain text if storage is unavailable
        cached = (booking.requirements_doc_text or '').strip()
        if cached:
            return JsonResponse({
                'doc_name': doc_name,
                'text':     cached,
                'blocks':   [{'type': 'paragraph', 'text': line}
                             for line in cached.split('\n') if line.strip()],
            })
        return JsonResponse({'error': f'Could not read file from storage: {str(e)}'}, status=500)

    try:
        result = _extract_docx_structured(raw)
    except Exception as e:
        print(f"[get_booking_doc_text] Parse failed booking={booking_id}: {e}\n{_tb.format_exc()}")
        return JsonResponse({'error': f'Could not parse document: {str(e)}'}, status=500)

    blocks     = result['blocks']
    plain_text = result['plain_text'] or '(No readable content found in this document)'

    # Persist/update plain text cache
    try:
        booking.requirements_doc_text = plain_text
        booking.save(update_fields=['requirements_doc_text'])
    except Exception as e:
        print(f"[get_booking_doc_text] Failed to persist text booking={booking_id}: {e}")

    return JsonResponse({'doc_name': doc_name, 'blocks': blocks, 'text': plain_text})


def download_booking_doc_as_pdf(request, booking_id):
    """
    Generates a beautifully formatted PDF from the booking requirements document.
    Always re-parses the .docx file so table structure is preserved exactly.
    Falls back to cached plain text only if storage is unreachable.
    Central admin and sub-admin only.
    """
    import io
    import traceback as _tb
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, HRFlowable,
        Table as RLTable, TableStyle as RLTableStyle, KeepTogether,
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return HttpResponse("Unauthorized", status=403)

    booking = get_object_or_404(RoomBooking, id=booking_id)

    # ── Always re-parse from the .docx to preserve table structure ─────────────
    blocks = None
    if booking.requirements_doc and booking.requirements_doc.name:
        try:
            from docx import Document  # noqa
        except ImportError:
            return HttpResponse("python-docx not installed. Run: pip install python-docx", status=500)
        try:
            _storage = booking.requirements_doc.storage
            with _storage.open(booking.requirements_doc.name, 'rb') as _f:
                raw = _f.read()
            result = _extract_docx_structured(raw)
            blocks = result['blocks']
            # Update plain text cache while we have the file
            try:
                _pt = result['plain_text'] or ''
                if _pt:
                    booking.requirements_doc_text = _pt
                    booking.save(update_fields=['requirements_doc_text'])
            except Exception:
                pass
        except Exception as e:
            print(f"[download_pdf] file read failed booking={booking_id}: {e}\n{_tb.format_exc()}")
            # Fall back to cached plain text
            cached = (booking.requirements_doc_text or '').strip()
            if cached:
                blocks = [{'type': 'paragraph', 'text': line}
                          for line in cached.split('\n') if line.strip()]
            else:
                return HttpResponse(f"Could not read document: {str(e)}", status=500)

    if not blocks:
        # No file attached and no cache
        return HttpResponse("No document content available for this booking.", status=404)

    # ── Colour palette ──────────────────────────────────────────────────────────
    C_INDIGO      = colors.HexColor('#4f46e5')   # header bg, table header
    C_INDIGO_DARK = colors.HexColor('#3730a3')   # gradient simulation
    C_INDIGO_LIGHT= colors.HexColor('#e0e7ff')   # section bg tint
    C_ROW_ALT     = colors.HexColor('#f0f4ff')   # table alt row
    C_BORDER      = colors.HexColor('#c7d2fe')   # table grid
    C_TEXT_DARK   = colors.HexColor('#1e293b')
    C_TEXT_MID    = colors.HexColor('#475569')
    C_TEXT_LIGHT  = colors.HexColor('#94a3b8')
    C_WHITE       = colors.white
    C_RULE        = colors.HexColor('#e2e8f0')

    # ── Page setup ──────────────────────────────────────────────────────────────
    PAGE_W = A4[0]
    MARGIN = 2.2 * cm
    USABLE = PAGE_W - 2 * MARGIN

    buffer  = io.BytesIO()
    pdf_doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN,  bottomMargin=MARGIN,
    )
    styles = getSampleStyleSheet()

    # ── Typography ──────────────────────────────────────────────────────────────
    s_title = ParagraphStyle(
        'Title', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=20,
        textColor=C_WHITE, leading=26, spaceAfter=0,
    )
    s_meta = ParagraphStyle(
        'Meta', parent=styles['Normal'],
        fontName='Helvetica', fontSize=9,
        textColor=colors.HexColor('#c7d2fe'), leading=14, spaceAfter=0,
    )
    s_label = ParagraphStyle(
        'Label', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8,
        textColor=C_TEXT_LIGHT, leading=11, spaceAfter=2,
        spaceBefore=0,
    )
    s_section = ParagraphStyle(
        'Section', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10,
        textColor=C_INDIGO, leading=14, spaceAfter=4, spaceBefore=10,
    )
    s_body = ParagraphStyle(
        'Body', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10,
        textColor=C_TEXT_DARK, leading=16, spaceAfter=5,
    )
    s_cell_hdr = ParagraphStyle(
        'CellHdr', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=9,
        textColor=C_WHITE, leading=13,
    )
    s_cell = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=9,
        textColor=C_TEXT_DARK, leading=13,
    )

    # ── Helpers ─────────────────────────────────────────────────────────────────
    def _esc(s):
        return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    try:
        start_str = timezone.localtime(booking.start_datetime).strftime('%d %b %Y, %H:%M')
        end_str   = timezone.localtime(booking.end_datetime).strftime('%d %b %Y, %H:%M')
    except Exception:
        start_str = str(booking.start_datetime)
        end_str   = str(booking.end_datetime)

    room_name    = booking.room.room_name if booking.room else '—'
    faculty_name = booking.faculty_name  or '—'
    faculty_email= booking.faculty_email or '—'

    # ── Header banner (indigo box with title + meta) ────────────────────────────
    HDR_W = USABLE
    hdr_data = [[
        Paragraph("Requirements Document", s_title),
        Paragraph(
            f"<b>Room</b>  {_esc(room_name)}<br/>"
            f"<b>Faculty</b>  {_esc(faculty_name)}<br/>"
            f"<b>Email</b>  {_esc(faculty_email)}<br/>"
            f"<b>Booking</b>  {_esc(start_str)}  →  {_esc(end_str)}",
            s_meta
        ),
    ]]
    hdr_tbl = RLTable(hdr_data, colWidths=[HDR_W*0.45, HDR_W*0.55])
    hdr_tbl.setStyle(RLTableStyle([
        ('BACKGROUND',   (0,0), (-1,-1), C_INDIGO),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,-1), 18),
        ('BOTTOMPADDING',(0,0), (-1,-1), 18),
        ('LEFTPADDING',  (0,0), (0,-1),  18),
        ('RIGHTPADDING', (-1,0),(-1,-1), 18),
        ('ROUNDEDCORNERS', [8, 8, 8, 8]),
    ]))

    elements = [hdr_tbl, Spacer(1, 0.5*cm)]

    # ── Purpose box ─────────────────────────────────────────────────────────────
    if booking.purpose:
        purpose_data = [[
            Paragraph("PURPOSE OF BOOKING", s_label),
            Paragraph(_esc(booking.purpose).replace('\n', '<br/>'), s_body),
        ]]
        purpose_tbl = RLTable(purpose_data, colWidths=[HDR_W*0.22, HDR_W*0.78])
        purpose_tbl.setStyle(RLTableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), C_INDIGO_LIGHT),
            ('LINEAFTER',     (0,0), (0,-1),  2, C_INDIGO),
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING',    (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('LEFTPADDING',   (0,0), (0,-1),  14),
            ('RIGHTPADDING',  (-1,0),(-1,-1), 14),
            ('LEFTPADDING',   (1,0), (1,-1),  14),
            ('ROUNDEDCORNERS', [6, 6, 6, 6]),
        ]))
        elements += [purpose_tbl, Spacer(1, 0.45*cm)]

    # ── Section heading ─────────────────────────────────────────────────────────
    elements.append(Paragraph("DOCUMENT CONTENT", s_label))
    elements.append(HRFlowable(width=USABLE, thickness=1.5, color=C_INDIGO, spaceAfter=8))

    # ── Content blocks ──────────────────────────────────────────────────────────
    for block in blocks:
        if block['type'] == 'paragraph':
            txt = _esc(block['text'])
            elements.append(Paragraph(txt, s_body))

        elif block['type'] == 'table':
            rows = block.get('rows', [])
            if not rows:
                continue

            num_cols = max(len(r) for r in rows)
            if num_cols == 0:
                continue

            # ── Smart column widths ─────────────────────────────────────────
            # Give each column equal width; if only 2 cols give first col 30%
            if num_cols == 2:
                col_widths = [USABLE * 0.28, USABLE * 0.72]
            elif num_cols == 3:
                col_widths = [USABLE * 0.25, USABLE * 0.40, USABLE * 0.35]
            else:
                col_widths = [USABLE / num_cols] * num_cols

            # ── Pad rows that have fewer cells than the max ─────────────────
            padded_rows = []
            for row in rows:
                padded = list(row) + [''] * (num_cols - len(row))
                padded_rows.append(padded)

            # ── First row = header; rest = data rows ────────────────────────
            header_row = [Paragraph(_esc(cell), s_cell_hdr) for cell in padded_rows[0]]
            data_rows  = [
                [Paragraph(_esc(cell), s_cell) for cell in row]
                for row in padded_rows[1:]
            ]
            table_data = [header_row] + data_rows

            tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)

            ts = RLTableStyle([
                # Header
                ('BACKGROUND',    (0,0), (-1,0),  C_INDIGO),
                ('TEXTCOLOR',     (0,0), (-1,0),  C_WHITE),
                ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0,0), (-1,0),  9),
                # Data rows — alternate shading
                ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
                ('FONTSIZE',      (0,1), (-1,-1), 9),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_WHITE, C_ROW_ALT]),
                # Grid
                ('GRID',          (0,0), (-1,-1), 0.5, C_BORDER),
                ('LINEBELOW',     (0,0), (-1,0),  1.5, C_INDIGO_DARK),
                # Padding
                ('VALIGN',        (0,0), (-1,-1), 'TOP'),
                ('TOPPADDING',    (0,0), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                ('LEFTPADDING',   (0,0), (-1,-1), 8),
                ('RIGHTPADDING',  (0,0), (-1,-1), 8),
            ])

            # If table has only 1 row (all header, no data), treat it as data
            if len(rows) == 1:
                ts.add('BACKGROUND', (0,0), (-1,0), C_ROW_ALT)
                ts.add('TEXTCOLOR',  (0,0), (-1,0), C_TEXT_DARK)
                ts.add('FONTNAME',   (0,0), (-1,0), 'Helvetica')

            tbl.setStyle(ts)
            # Keep table header + first few data rows together across page breaks
            elements.append(KeepTogether([tbl]))
            elements.append(Spacer(1, 0.35*cm))

    # ── Footer rule ─────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width=USABLE, thickness=0.5, color=C_RULE))
    elements.append(Spacer(1, 0.1*cm))
    elements.append(Paragraph(
        f"Generated by AURA · {booking.room.room_name if booking.room else ''} · {start_str}",
        ParagraphStyle('Footer', parent=styles['Normal'],
                       fontName='Helvetica', fontSize=7, textColor=C_TEXT_LIGHT,
                       alignment=TA_CENTER)
    ))

    pdf_doc.build(elements)
    pdf_data = buffer.getvalue()
    buffer.close()

    safe_name = f"booking_{booking_id}_requirements.pdf"
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response


def aura_generate_report_excel(request):
    """
    Generate Excel report for AURA data - matches dashboard display format.
    """
    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return HttpResponse("Unauthorized", status=403)

    model_name = request.GET.get('model')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    model_map = {
        'issues': Issue, 'items': Item, 'rooms': Room, 
        'bookings': RoomBooking, 'purchases': Purchase,
        'vendors': Vendor, 'departments': Department
    }
    
    model = model_map.get(model_name)
    if not model:
        return HttpResponse("Invalid Model", status=400)
    
    qs = model.objects.all().order_by('-id')
    if date_from and date_to:
        if model_name == 'bookings':
            qs = qs.filter(start_datetime__date__range=[date_from, date_to])
        elif model_name in ['issues', 'items', 'rooms', 'purchases']:
            qs = qs.filter(created_on__date__range=[date_from, date_to])
    
    # Create Excel file
    output = io.BytesIO()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"AURA {model_name.capitalize()}"
    
    # Dynamic headers based on model type
    if model_name == 'rooms':
        headers = ["ID", "Room", "Metadata"]
    elif model_name == 'bookings':
        headers = ["ID", "Room Booked", "Metadata"]
    elif model_name == 'issues':
        headers = ["ID", "Issue", "Metadata"]
    elif model_name == 'items':
        headers = ["ID", "Items", "Metadata"]
    elif model_name == 'purchases':
        headers = ["ID", "Purchase ID", "Metadata"]
    elif model_name == 'vendors':
        headers = ["ID", "Vendors", "Metadata"]
    elif model_name == 'departments':
        headers = ["ID", "Department/Cell/Office", "Metadata"]
    else:
        headers = ["ID", "Record", "Details"]
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data - EXACT SAME LOGIC AS aura_data_manager
    for obj in qs:
        label = str(obj)
        detail = "General Record"
        
        try:
            if model_name == 'rooms':
                label = f"{obj.room_name}\nIncharge: {obj.incharge}"
                detail = f"Category: {obj.get_room_category_display()} | Capacity: {obj.capacity}"
                
            elif model_name == 'bookings':
                label = f"{obj.faculty_name}\n{obj.faculty_email}"
                start_local = timezone.localtime(obj.start_datetime)
                end_local = timezone.localtime(obj.end_datetime)
                detail = f"Room: {obj.room.room_name}\nDate: {start_local.strftime('%d %b, %Y')}\nTime: {start_local.strftime('%H:%M')} - {end_local.strftime('%H:%M')}"
                
            elif model_name == 'issues':
                label = obj.subject
                assigned = obj.assigned_to.user.get_full_name() if obj.assigned_to else "N/A"
                detail = f"Email: {obj.reporter_email}\nTicket ID: {obj.ticket_id}\nStatus: {obj.status}\nRoom: {obj.room.room_name}\nAssigned: {assigned}"
                
            elif model_name == 'items':
                from inventory.models import AssetTag
                label = obj.item_name
                room_name = obj.room.room_name if obj.room else 'Master Inventory'
                prod_code = obj.product_code or '—'
                tag_range = '—'
                if obj.room and obj.product_code:
                    tags = AssetTag.objects.filter(
                        item_name=obj.item_name,
                        assigned_room=obj.room
                    ).order_by('tag_id')
                    if tags.exists():
                        tag_range = f"{tags.first().tag_id} → {tags.last().tag_id}"
                detail = f"Room: {room_name}\nProduct Code: {prod_code}\nAsset Tags: {tag_range}\nQty: {obj.total_count}\nAvailable: {obj.available_count}\nIn Use: {obj.in_use}"

            elif model_name == 'purchases':
                label = f"{obj.purchase_id or 'Pending ID'}<br/>Room: {obj.room.room_name if obj.room else 'No Room'}"
                item_name = obj.item.item_name if obj.item else 'N/A'
                vendor_name = obj.vendor.vendor_name if obj.vendor else 'No Vendor'
                detail = f"Item: {item_name}<br/>Vendor: {vendor_name}<br/>Status: {obj.status.title()}"
                
            elif model_name == 'vendors':
                label = obj.vendor_name
                detail = f"Email: {obj.email}\nContact: {obj.contact_number}"
                
            elif model_name == 'departments':
                label = obj.department_name
                room_count = Room.objects.filter(department=obj).count()
                detail = f"Total Rooms: {room_count}"
                
        except Exception as e:
            detail = f"Data Error: {str(e)}"
        
        # Append row — uniform 3 columns for all modules
        ws.append([obj.id, label, detail])
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 40  # Label
    ws.column_dimensions['C'].width = 60  # Details / Metadata
    
    # Set row heights for better readability
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 60
    
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Blixtro_AURA_{model_name}_Report.xlsx"'
    return response


class MasterInventoryImportView(LoginRequiredMixin, CentralAdminRequiredMixin, FormView):
    """
    Master Inventory Upload - No room context, items stay unassigned
    """
    template_name = 'central_admin/master_inventory_import_upload.html'
    form_class = ExcelUploadForm

    def get_success_url(self):
        return reverse('central_admin:master_inventory_import')

    def form_valid(self, form):
        upload_file = form.cleaned_data['file']
        try:
            excel = pd.ExcelFile(upload_file)
        except Exception as e:
            form.add_error('file', f'Invalid Excel file: {e}')
            return self.form_invalid(form)

        org = self.request.user.profile.org
        preview = {'items': [], 'errors': []}

        if 'Items' not in excel.sheet_names:
            preview['errors'].append("Sheet named 'Items' not found. Please name your sheet 'Items'.")
        else:
            df = excel.parse('Items')
            df.columns = [str(c).strip() for c in df.columns]

            # Only Item Name is strictly required as a column
            mandatory = ["Item Name"]
            missing = [c for c in mandatory if c not in df.columns]

            if missing:
                preview['errors'].append(f"Excel is missing mandatory columns: {missing}")
            else:
                for idx, row in df.iterrows():
                    rownum = idx + 2
                    row_errors = []

                    # ── Item Name (required) ──────────────────────────────────
                    name = str(row.get('Item Name', '')).strip()
                    if not name or name.lower() == 'nan':
                        row_errors.append('Item Name is required and cannot be empty.')
                        name = '—'

                    # ── Category (optional) ───────────────────────────────────
                    cat_raw = row.get('Category', None)
                    cat_str = str(cat_raw).strip() if cat_raw is not None else ''
                    cat_display = cat_str if (cat_str and cat_str.lower() != 'nan') else '-'

                    # ── Brand (optional) ──────────────────────────────────────
                    brand_raw = row.get('Brand', None)
                    brand_str = str(brand_raw).strip() if brand_raw is not None else ''
                    brand_display = brand_str if (brand_str and brand_str.lower() != 'nan') else '-'

                    # ── Total Count (optional) ────────────────────────────────
                    tc_raw = row.get('Total Count', None)
                    if tc_raw is None or str(tc_raw).strip() == '' or str(tc_raw).lower() == 'nan':
                        total_count_display = '-'
                    else:
                        try:
                            total_count_display = int(tc_raw)
                        except (ValueError, TypeError):
                            total_count_display = '-'
                            row_errors.append('Total Count must be a whole number if provided.')

                    # ── Cost (optional) ───────────────────────────────────────
                    cost_raw = row.get('Cost', None)
                    if cost_raw is None or str(cost_raw).strip() == '' or str(cost_raw).lower() == 'nan':
                        cost_display = '-'
                    else:
                        try:
                            from decimal import Decimal, InvalidOperation
                            cost_display = Decimal(str(cost_raw))
                        except (ValueError, TypeError, InvalidOperation):
                            cost_display = '-'
                            row_errors.append('Cost must be a valid decimal number if provided.')

                    preview['items'].append({
                        'rownum': rownum,
                        'name': name,
                        'category': cat_display,
                        'brand': brand_display,
                        'total_count': total_count_display,
                        'cost': cost_display,
                        'errors': row_errors
                    })

        has_errors = bool(preview['errors'] or any(r['errors'] for r in preview['items']))

        return render(self.request, "central_admin/master_inventory_import_view.html", {
            "preview": preview,
            "has_errors": has_errors,
        })


class MasterInventoryImportConfirmView(LoginRequiredMixin, CentralAdminRequiredMixin, View):
    """
    Process Master Inventory Import - Creates UNASSIGNED items (no room link)
    """
    def post(self, request, *args, **kwargs):
        upload_file = request.FILES.get('file')
        if not upload_file:
            messages.error(request, "No file uploaded for confirmation.")
            return redirect('central_admin:master_inventory_import')

        org = request.user.profile.org
        
        try:
            import pandas as pd
            df = pd.read_excel(upload_file, sheet_name='Items')
        except Exception as e:
            messages.error(request, f"Processing error: {e}")
            return redirect('central_admin:master_inventory_import')

        import_count = 0
        from decimal import Decimal
        from inventory.models import Category, Brand

        for _, row in df.iterrows():
            name = str(row.get('Item Name', '')).strip()
            if not name or name.lower() == 'nan':
                continue

            # ── Category (optional — fall back to 'Uncategorised') ────────────
            cat_raw = row.get('Category', None)
            cat_name = str(cat_raw).strip() if cat_raw is not None else ''
            if cat_name and cat_name.lower() != 'nan':
                category, _ = Category.objects.get_or_create(
                    organisation=org, room=None, category_name=cat_name)
            else:
                cat_name = 'Uncategorised'
                category, _ = Category.objects.get_or_create(
                    organisation=org, room=None, category_name='Uncategorised')

            # ── Brand (optional — fall back to 'Unknown') ─────────────────────
            brand_raw = row.get('Brand', None)
            brand_name = str(brand_raw).strip() if brand_raw is not None else ''
            if brand_name and brand_name.lower() != 'nan':
                brand, _ = Brand.objects.get_or_create(
                    organisation=org, room=None, brand_name=brand_name)
            else:
                brand_name = 'Unknown'
                brand, _ = Brand.objects.get_or_create(
                    organisation=org, room=None, brand_name='Unknown')

            # ── Total Count (optional — fall back to 0) ───────────────────────
            tc_raw = row.get('Total Count', None)
            if tc_raw is None or str(tc_raw).strip() == '' or str(tc_raw).lower() == 'nan':
                total_count = 0
            else:
                try:
                    total_count = int(tc_raw)
                except (ValueError, TypeError):
                    total_count = 0

            # ── Cost (optional — stored as None if absent) ────────────────────
            cost_raw = row.get('Cost', None)
            try:
                cost = Decimal(str(cost_raw)) if cost_raw is not None and str(cost_raw).lower() not in ('nan', '') else None
            except Exception:
                cost = None

            # ── Create / update unassigned master item ────────────────────────
            Item.objects.update_or_create(
                organisation=org,
                room=None,  # Master inventory - unassigned
                item_name=name,
                defaults={
                    'category': category,
                    'brand': brand,
                    'total_count': total_count,
                    'cost': cost,
                    'is_listed': True,
                    'item_description': f"{brand_name} {name} - Master Inventory"
                }
            )
            import_count += 1

        messages.success(request, f"Successfully imported {import_count} items to Master Inventory.")
        return redirect("central_admin:aura_dashboard")
    
class MasterInventoryListView(LoginRequiredMixin, CentralAdminRequiredMixin, TemplateView):
    """
    View all unassigned items in Master Inventory
    """
    template_name = 'central_admin/master_inventory_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.profile.org

        master_items = Item.objects.filter(
            organisation=org,
            room__isnull=True,
            is_listed=True
        ).select_related('category', 'brand').order_by('item_name')

        # ── Bulk aggregations (4 queries total, no per-item DB hits) ──
        assigned_map = {
            row['item_name']: row['total'] or 0
            for row in Item.objects.filter(organisation=org, room__isnull=False)
            .values('item_name').annotate(total=models.Sum('total_count'))
        }
        archived_map = {
            row['item_name']: row['total'] or 0
            for row in Item.objects.filter(organisation=org)
            .values('item_name').annotate(total=models.Sum('archived_count'))
        }
        sc_map = {}
        for row in (
            SC.objects.filter(
                component_item__organisation=org,
                component_item__room__isnull=False,
                status__in=['inactive', 'under_maintenance', 'disposed'],
            )
            .values('component_item__item_name', 'status')
            .annotate(cnt=models.Count('id'))
        ):
            sc_map.setdefault(row['component_item__item_name'], {})[row['status']] = row['cnt']

        items_data = []
        for item in master_items:
            name = item.item_name
            assigned = assigned_map.get(name, 0)
            available = item.total_count
            total_items = available + assigned
            cpu = item.cost or 0
            sc = sc_map.get(name, {})
            items_data.append({
                'item': item,
                'available_stock': available,
                'assigned_count': assigned,
                'total_items': total_items,
                'cpu': cpu,
                'total_cost': round(float(cpu) * total_items, 2) if cpu else 0,
                'archived_count': archived_map.get(name, 0),
                'inactive_count': sc.get('inactive', 0),
                'under_maintenance_count': sc.get('under_maintenance', 0),
                'disposed_count': sc.get('disposed', 0),
            })

        context['items_data'] = items_data
        context['total_items'] = len(items_data)
        return context
    
def master_inventory_export_pdf(request):
    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return HttpResponse("Unauthorized", status=403)

    org = profile.org
    fields = request.GET.getlist('fields')
    if not fields:
        fields = ['name', 'category', 'brand', 'assigned', 'stock', 'cost']

    master_items = Item.objects.filter(
        organisation=org,
        room__isnull=True,
        is_listed=True
    ).select_related('category', 'brand').order_by('item_name')

    # ── Bulk aggregations — same pattern as MasterInventoryListView ──
    assigned_map = {
        row['item_name']: row['total'] or 0
        for row in Item.objects.filter(organisation=org, room__isnull=False)
        .values('item_name').annotate(total=models.Sum('total_count'))
    }
    archived_map = {
        row['item_name']: row['total'] or 0
        for row in Item.objects.filter(organisation=org)
        .values('item_name').annotate(total=models.Sum('archived_count'))
    }
    sc_map = {}
    for row in (
        SC.objects.filter(
            component_item__organisation=org,
            component_item__room__isnull=False,
            status__in=['inactive', 'under_maintenance', 'disposed'],
        )
        .values('component_item__item_name', 'status')
        .annotate(cnt=models.Count('id'))
    ):
        sc_map.setdefault(row['component_item__item_name'], {})[row['status']] = row['cnt']

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=20, leftMargin=20,
                            topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Master Inventory Report", styles['Title']))
    elements.append(Spacer(1, 10))

    text_col_map = {
        'prodcode': ('Product Code', 70),
        'name':     ('Item Name',  80),
        'category': ('Category',   70),
        'brand':    ('Brand',      70),
    }
    fixed_cols = [
        ('Assigned\nto Rooms',    55),
        ('Available\nin Stock',   55),
        ('Total\nItems',          55),
        ('Cost /\nUnit (₹)',      60),
        ('Total\nCost (₹)',       65),
        ('Archived',              50),
        ('Inactive\nComponents',  55),
        ('Under\nMaintenance',    55),
        ('Disposed\nComponents',  55),
    ]

    selected_text = [k for k in ['prodcode', 'name', 'category', 'brand'] if k in fields]
    from reportlab.lib.styles import ParagraphStyle
    header_style = ParagraphStyle('header', parent=styles['Normal'], textColor=colors.whitesmoke, fontName='Helvetica-Bold')
    headers = [Paragraph(text_col_map[k][0], header_style) for k in selected_text]
    col_widths = [text_col_map[k][1] for k in selected_text]

    for label, w in fixed_cols:
        headers.append(Paragraph(label, header_style))
        col_widths.append(w)

    table_data = [headers]

    for item in master_items:
        name = item.item_name
        assigned = assigned_map.get(name, 0)
        available = item.total_count
        total_items = available + assigned
        cpu = item.cost or 0
        total_cost = (cpu * total_items) if cpu else 0
        archived_c = archived_map.get(name, 0)
        sc = sc_map.get(name, {})
        inactive_c = sc.get('inactive', 0)
        under_maintenance_c = sc.get('under_maintenance', 0)
        disposed_c = sc.get('disposed', 0)

        row = []
        if 'prodcode' in selected_text:
            row.append(Paragraph(item.product_code or '—', styles['Normal']))
        if 'name' in selected_text:
            row.append(Paragraph(item.item_name, styles['Normal']))
        if 'category' in selected_text:
            row.append(Paragraph(item.category.category_name, styles['Normal']))
        if 'brand' in selected_text:
            row.append(Paragraph(item.brand.brand_name, styles['Normal']))

        row += [
            Paragraph(str(assigned), styles['Normal']),
            Paragraph(str(available), styles['Normal']),
            Paragraph(str(total_items), styles['Normal']),
            Paragraph(f"₹{cpu}", styles['Normal']),
            Paragraph(f"₹{total_cost:.2f}", styles['Normal']),
            Paragraph(str(archived_c), styles['Normal']),
            Paragraph(str(inactive_c), styles['Normal']),
            Paragraph(str(under_maintenance_c), styles['Normal']),
            Paragraph(str(disposed_c), styles['Normal']),
        ]
        table_data.append(row)

    if len(table_data) == 1:
        elements.append(Paragraph("No items found.", styles['Normal']))
    else:
        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (len(selected_text), 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Master_Inventory.pdf"'
    return response

def master_inventory_export_excel(request):
    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return HttpResponse("Unauthorized", status=403)

    org = profile.org
    fields = request.GET.getlist('fields')
    if not fields:
        fields = ['name', 'category', 'brand', 'stock', 'cost']

    master_items = Item.objects.filter(
        organisation=org,
        room__isnull=True,
        is_listed=True
    ).select_related('category', 'brand').order_by('item_name')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Inventory"

    text_col_map = {
        'prodcode': 'Product Code',
        'name':     'Item Name',
        'category': 'Category',
        'brand':    'Brand',
    }
    selected_text = [k for k in ['prodcode', 'name', 'category', 'brand'] if k in fields]
    headers = [text_col_map[k] for k in selected_text]

    headers += [
        'Assigned to Rooms',
        'Available in Stock',
        'Total Items',
        'Cost per Unit (₹)',
        'Total Cost (₹)',
        'Archived',
        'Inactive Components',
        'Under Maintenance',
        'Disposed Components',
    ]

    ws.append(headers)

    dark = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # ── Bulk aggregations — no per-item DB hits ──
    assigned_map_xl = {
        row['item_name']: row['total'] or 0
        for row in Item.objects.filter(organisation=org, room__isnull=False)
        .values('item_name').annotate(total=models.Sum('total_count'))
    }
    archived_map_xl = {
        row['item_name']: row['total'] or 0
        for row in Item.objects.filter(organisation=org)
        .values('item_name').annotate(total=models.Sum('archived_count'))
    }
    sc_map_xl = {}
    for row in (
        SC.objects.filter(
            component_item__organisation=org,
            component_item__room__isnull=False,
            status__in=['inactive', 'under_maintenance', 'disposed'],
        )
        .values('component_item__item_name', 'status')
        .annotate(cnt=models.Count('id'))
    ):
        sc_map_xl.setdefault(row['component_item__item_name'], {})[row['status']] = row['cnt']

    for item in master_items:
        name = item.item_name
        assigned = assigned_map_xl.get(name, 0)
        available = item.total_count
        total_items = available + assigned
        cpu = float(item.cost) if item.cost else 0
        total_cost = round(cpu * total_items, 2)
        archived_xl = archived_map_xl.get(name, 0)
        sc = sc_map_xl.get(name, {})
        inactive_xl = sc.get('inactive', 0)
        under_maintenance_xl = sc.get('under_maintenance', 0)
        disposed_xl = sc.get('disposed', 0)

        row = []
        if 'prodcode' in selected_text:    row.append(item.product_code or '—')
        if 'name' in selected_text:        row.append(item.item_name)
        if 'category' in selected_text:    row.append(item.category.category_name)
        if 'brand' in selected_text:       row.append(item.brand.brand_name)

        row += [assigned, available, total_items, cpu, total_cost, archived_xl,
                inactive_xl, under_maintenance_xl, disposed_xl]
        ws.append(row)

    thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") \
               if i % 2 == 0 else None
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border = thin
            if fill:
                cell.fill = fill

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=len(selected_text)):
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="left")

    text_widths = {'prodcode': 16, 'name': 32, 'category': 20, 'brand': 18}
    fixed_widths = [18, 18, 14, 18, 16, 14, 18, 18, 18]

    col = 1
    for k in selected_text:
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = text_widths[k]
        col += 1
    for w in fixed_widths:
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        col += 1

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Master_Inventory.xlsx"'
    return response

class AssignInventoryView(LoginRequiredMixin, CentralAdminRequiredMixin, TemplateView):
    template_name = 'central_admin/assign_inventory.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.profile.org
        master_items = Item.objects.filter(
            organisation=org,
            room__isnull=True,
            is_listed=True
        ).select_related('category', 'brand').order_by('item_name')
        context['master_items'] = master_items
        context['room_categories'] = Room.ROOM_CATEGORIES
        return context

def assign_inventory_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    import json
    from django.db import transaction

    try:
        data = json.loads(request.body)
        print("DEBUG DATA:", data)
        print("DEBUG item_id:", data.get('item_id'))
        print("DEBUG room_ids:", data.get('room_ids'))
        print("DEBUG quantity:", data.get('quantity'))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    master_item_id = data.get('item_id')
    room_ids = data.get('room_ids', [])
    quantity = data.get('quantity')

    if not master_item_id or not room_ids or not quantity:
        return JsonResponse({'error': 'item_id, room_ids and quantity are required.'}, status=400)

    if not isinstance(room_ids, list) or len(room_ids) == 0:
        return JsonResponse({'error': 'room_ids must be a non-empty list.'}, status=400)

    try:
        quantity = int(quantity)
        if quantity <= 0:
            raise ValueError()
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Quantity must be a positive integer.'}, status=400)

    # Convert room_ids to integers
    try:
        room_ids = [int(rid) for rid in room_ids]
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid room_ids.'}, status=400)

    org = profile.org

    try:
        master_item = Item.objects.get(id=master_item_id, organisation=org, room__isnull=True)
    except Item.DoesNotExist:
        return JsonResponse({'error': 'Master inventory item not found.'}, status=404)

    total_needed = quantity * len(room_ids)
    if total_needed > master_item.total_count:
        return JsonResponse({
            'error': f'Insufficient stock. Need {total_needed} units for {len(room_ids)} rooms but only {master_item.total_count} available.'
        }, status=400)

    from inventory.models import Category, Brand

    with transaction.atomic():
        master_item.total_count -= total_needed
        master_item.save()

        for room_id in room_ids:
            try:
                room = Room.objects.get(id=room_id, organisation=org)
            except Room.DoesNotExist:
                continue

            room_category, _ = Category.objects.get_or_create(
                organisation=org,
                room=room,
                category_name=master_item.category.category_name
            )

            room_brand, _ = Brand.objects.get_or_create(
                organisation=org,
                room=room,
                brand_name=master_item.brand.brand_name
            )

            room_item = Item.objects.filter(
                organisation=org,
                room=room,
                item_name=master_item.item_name,
            ).first()

            if room_item:
                room_item.total_count += quantity
                room_item.save()
            else:
                Item.objects.create(
                    organisation=org,
                    room=room,
                    item_name=master_item.item_name,
                    category=room_category,
                    brand=room_brand,
                    total_count=quantity,
                    cost=master_item.cost,
                    is_listed=True,
                    item_description=master_item.item_description,
                    created_by=profile,
                )

    # Assign asset tags to rooms after inventory assignment
    from inventory.models import AssetTag
    if Item.objects.filter(organisation=org, item_name=master_item.item_name, room__isnull=True, is_listed=True).filter(product_code__isnull=False).exists():
        _assign_tags_to_rooms(org, master_item.item_name)
    return JsonResponse({
        'success': True,
        'message': f'Assigned {quantity} unit(s) of "{master_item.item_name}" to {len(room_ids)} room(s). Total deducted: {total_needed}.',
        'master_remaining': master_item.total_count,
    })

def unassign_inventory_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    room_item_id = data.get('room_item_id')
    quantity     = data.get('quantity')

    if not room_item_id or not quantity:
        return JsonResponse({'error': 'room_item_id and quantity are required.'}, status=400)

    try:
        quantity = int(quantity)
        if quantity <= 0:
            raise ValueError()
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Quantity must be a positive integer.'}, status=400)

    org = profile.org

    try:
        room_item = Item.objects.get(id=room_item_id, organisation=org, room__isnull=False)
    except Item.DoesNotExist:
        return JsonResponse({'error': 'Room item not found.'}, status=404)

    if quantity > room_item.total_count:
        return JsonResponse({
            'error': f'Only {room_item.total_count} units assigned. Cannot return more.'
        }, status=400)

    from django.db import transaction
    with transaction.atomic():
        # Find master item
        master_item = Item.objects.filter(
            organisation=org,
            room__isnull=True,
            item_name=room_item.item_name,
            is_listed=True,
        ).first()

        # Deduct from room
        room_item.total_count -= quantity
        if room_item.total_count == 0:
            room_item.delete()
        else:
            room_item.save(update_fields=['total_count', 'updated_on'])

        # Return to master
        if master_item:
            master_item.total_count += quantity
            master_item.save(update_fields=['total_count', 'updated_on'])
        else:
            Item.objects.create(
                organisation=org,
                room=None,
                item_name=room_item.item_name,
                category=room_item.category,
                brand=room_item.brand,
                total_count=quantity,
                cost=room_item.cost,
                is_listed=True,
                item_description=room_item.item_description or room_item.item_name,
                created_by=profile,
            )

    return JsonResponse({
        'success': True,
        'message': f'{quantity} unit(s) of "{room_item.item_name}" returned to master stock.',
    })

def get_master_items_api(request):
    profile = request.user.profile
    if not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    org = profile.org
    items = Item.objects.filter(
        organisation=org,
        room__isnull=True,
        is_listed=True,
    ).select_related('category', 'brand').order_by('item_name')

    return JsonResponse({'items': [{
        'id': item.id,
        'name': item.item_name,
        'category': item.category.category_name,
        'brand': item.brand.brand_name,
        'available_stock': item.total_count,
    } for item in items]})

def save_product_code(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    import json
    from inventory.models import AssetTag
    data = json.loads(request.body)
    item_id = data.get('item_id')
    code = (data.get('product_code') or '').strip()

    if not item_id:
        return JsonResponse({'error': 'item_id required'}, status=400)
    if code and not _re.match(r'^[\w\-\.@#&*()]{1,12}$', code):
        return JsonResponse({'error': 'Invalid code. Max 12 chars, alphanumeric + special chars only.'}, status=400)

    org = profile.org
    master_item = get_object_or_404(Item, id=item_id, organisation=org, room__isnull=True)
    old_code = master_item.product_code or ''
    master_item.product_code = code or None
    master_item.save(update_fields=['product_code', 'updated_on'])

    # Propagate to all room items with same item_name
    Item.objects.filter(
        organisation=org,
        item_name=master_item.item_name,
        room__isnull=False
    ).update(product_code=master_item.product_code)

    # Auto-generate asset tags if code provided
    tags_created = 0
    if code:
        # Count total items across master + all rooms
        total = Item.objects.filter(
            organisation=org,
            item_name=master_item.item_name,
        ).aggregate(t=models.Sum('total_count'))['t'] or 0

        existing_count = AssetTag.objects.filter(
            organisation=org,
            item_name=master_item.item_name,
        ).count()

        # Only generate new tags beyond existing ones
        for i in range(existing_count + 1, total + 1):
            tag_id = f"{code}-{i:02d}"
            AssetTag.objects.get_or_create(
                organisation=org,
                item_name=master_item.item_name,
                tag_id=tag_id,
                defaults={'assigned_room': None}
            )
            tags_created += 1

        # Assign tags to rooms in order
        _assign_tags_to_rooms(org, master_item.item_name)

    return JsonResponse({
        'success': True,
        'product_code': code,
        'tags_created': tags_created,
    })


def _assign_tags_to_rooms(org, item_name):
    """Assign unassigned asset tags to rooms in order based on room item counts."""
    from inventory.models import AssetTag
    unassigned = list(AssetTag.objects.filter(
        organisation=org,
        item_name=item_name,
        assigned_room__isnull=True
    ).order_by('tag_id'))

    room_items = Item.objects.filter(
        organisation=org,
        item_name=item_name,
        room__isnull=False
    ).select_related('room').order_by('room__room_name')

    idx = 0
    for room_item in room_items:
        needed = room_item.total_count
        already = AssetTag.objects.filter(
            organisation=org,
            item_name=item_name,
            assigned_room=room_item.room
        ).count()
        to_assign = needed - already
        for _ in range(to_assign):
            if idx >= len(unassigned):
                break
            tag = unassigned[idx]
            tag.assigned_room = room_item.room
            tag.save(update_fields=['assigned_room'])
            idx += 1


def save_item_edit(request):
    """Inline edit for category, brand, cost on master inventory."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    import json
    from decimal import Decimal, InvalidOperation
    data = json.loads(request.body)
    item_id = data.get('item_id')
    org = profile.org

    master_item = get_object_or_404(Item, id=item_id, organisation=org, room__isnull=True)

    # Category
    cat_name = (data.get('category') or '').strip()
    if cat_name:
        from inventory.models import Category as Cat
        cat, _ = Cat.objects.get_or_create(organisation=org, room=None, category_name=cat_name)
        master_item.category = cat

    # Brand
    brand_name = (data.get('brand') or '').strip()
    if brand_name:
        from inventory.models import Brand as Br
        br, _ = Br.objects.get_or_create(organisation=org, room=None, brand_name=brand_name)
        master_item.brand = br

    # Cost
    cost_val = data.get('cost')
    if cost_val is not None and str(cost_val).strip() != '':
        try:
            master_item.cost = Decimal(str(cost_val))
        except InvalidOperation:
            return JsonResponse({'error': 'Invalid cost value'}, status=400)
    elif cost_val == '':
        master_item.cost = None

    master_item.save()
    return JsonResponse({'success': True})


def get_asset_tags(request):
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from inventory.models import AssetTag
    item_name = request.GET.get('item_name', '')
    org = profile.org

    tags = AssetTag.objects.filter(
        organisation=org,
        item_name=item_name,
    ).select_related('assigned_room').order_by('tag_id')

    return JsonResponse({'tags': [{
        'tag_id': t.tag_id,
        'assigned_room': t.assigned_room.room_name if t.assigned_room else '—',
    } for t in tags]})


def get_room_asset_tags(request):
    """For room incharge — get asset tags assigned to their room for an item."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from inventory.models import AssetTag
    item_name = request.GET.get('item_name', '')
    room_slug = request.GET.get('room_slug', '')

    tags = AssetTag.objects.filter(
        item_name=item_name,
        assigned_room__slug=room_slug,
    ).order_by('tag_id')

    return JsonResponse({'tags': [{'tag_id': t.tag_id} for t in tags]})

# ─────────────────────────────────────────────────────────────────────────────
# FORWARD BOOKING REQUIREMENTS  (AURA — Confirmed Booking Files)
# ─────────────────────────────────────────────────────────────────────────────

@require_POST
def forward_booking_requirements(request, booking_id):
    """
    Receives a JSON payload from the Forward overlay and sends a
    professional HTML email to the specified recipient.

    Payload  (application/json):
    {
        "email":      "recipient@sfscollege.in",
        "department": "Tech",
        "blocks": [
            {"type": "paragraph", "text": "..."},
            {"type": "table",     "rows": [["h1","h2"], ["v1","v2"]]},
            ...
        ]
    }

    Returns:
        {"status": "success"}   on success
        {"error":  "..."}       on failure
    """
    from django.core.mail import EmailMultiAlternatives
    from django.conf import settings as _s
    import re as _re

    # ── Auth ─────────────────────────────────────────────────────────────────
    profile = getattr(request.user, 'profile', None)
    if not profile or not (profile.is_central_admin or profile.is_sub_admin):
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    # ── Parse body ────────────────────────────────────────────────────────────
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON body.'}, status=400)

    email      = (data.get('email') or '').strip().lower()
    department = (data.get('department') or '').strip()
    blocks     = data.get('blocks', [])

    # ── Validate email ────────────────────────────────────────────────────────
    if not email:
        return JsonResponse({'error': 'Recipient email is required.'}, status=400)
    if not email.endswith('@sfscollege.in'):
        return JsonResponse({'error': 'Only @sfscollege.in email addresses are allowed.'}, status=400)
    email_re = _re.compile(r'^[^\s@]+@sfscollege\.in$', _re.IGNORECASE)
    if not email_re.match(email):
        return JsonResponse({'error': 'Please provide a valid @sfscollege.in email address.'}, status=400)
    if not department:
        return JsonResponse({'error': 'Department is required.'}, status=400)
    if not blocks:
        return JsonResponse({'error': 'No content blocks selected.'}, status=400)

    # ── Fetch booking ─────────────────────────────────────────────────────────
    booking = get_object_or_404(RoomBooking, id=booking_id)

    try:
        from django.utils import timezone as _tz
        start_local = _tz.localtime(booking.start_datetime).strftime('%d %B %Y, %I:%M %p')
        end_local   = _tz.localtime(booking.end_datetime).strftime('%I:%M %p')
    except Exception:
        start_local = str(booking.start_datetime)
        end_local   = str(booking.end_datetime)

    room_name    = booking.room.room_name if booking.room else '—'
    faculty_name = booking.faculty_name   or '—'
    faculty_email= booking.faculty_email  or '—'
    purpose      = booking.purpose        or '—'
    sender_name  = str(profile)
    sender_role  = 'Central Admin' if (profile.is_central_admin and not profile.is_sub_admin) else 'Sub Admin'

    subject = f'Room Booking Requirements — {room_name} | {start_local}'

    # ── Build HTML content blocks ─────────────────────────────────────────────
    def _esc(s):
        return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

    content_html_parts = []
    for blk in blocks:
        btype = blk.get('type', 'paragraph')
        if btype == 'paragraph':
            text = blk.get('text', '')
            if text:
                content_html_parts.append(
                    f'<p style="margin:0 0 10px;color:#334155;line-height:1.7;">{_esc(text)}</p>'
                )
        elif btype == 'table':
            rows = blk.get('rows', [])
            if rows:
                table_html = (
                    '<div style="overflow-x:auto;margin-bottom:14px;border-radius:8px;border:1px solid #e2e8f0;">'
                    '<table style="min-width:100%;width:auto;border-collapse:collapse;font-size:0.85rem;">'
                )
                for ri, row in enumerate(rows):
                    tag = 'th' if ri == 0 else 'td'
                    bg  = '#0f172a' if ri == 0 else ('#f8fafc' if ri % 2 == 0 else '#ffffff')
                    col = '#ffffff' if ri == 0 else '#334155'
                    fw  = '700' if ri == 0 else '400'
                    table_html += '<tr>'
                    for cell in row:
                        table_html += (
                            f'<{tag} style="padding:8px 12px;border-bottom:1px solid #e2e8f0;'
                            f'border-right:1px solid #e2e8f0;background:{bg};color:{col};'
                            f'font-weight:{fw};text-align:left;white-space:nowrap;">'
                            f'{_esc(cell)}</{tag}>'
                        )
                    table_html += '</tr>'
                table_html += '</table></div>'
                content_html_parts.append(table_html)

    content_html = '\n'.join(content_html_parts) or '<p style="color:#94a3b8;">No content provided.</p>'

    # ── Full professional HTML email ───────────────────────────────────────────
    html_body = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{_esc(subject)}</title>
</head>
<body style="margin:0;padding:0;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:40px 20px;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0"
             style="background:#ffffff;border-radius:16px;overflow:hidden;
                    box-shadow:0 4px 24px rgba(0,0,0,0.07);max-width:600px;">

        <!-- Header -->
        <tr>
          <td style="background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 100%);
                     padding:32px 36px;text-align:center;">

            <!-- Pure-CSS logo mark: works in Gmail, Outlook, Apple Mail -->
            <table cellpadding="0" cellspacing="0" style="margin:0 auto 10px;">
              <tr>
                <!-- Lightning bolt shape via border-trick polygon -->
                <td style="padding:0 12px 0 0;vertical-align:middle;">
                  <table cellpadding="0" cellspacing="0">
                    <tr>
                      <!-- Outer dark-blue badge -->
                      <td style="width:46px;height:46px;background:#013198;border-radius:10px;
                                 text-align:center;vertical-align:middle;line-height:1;">
                        <!-- Inner yellow lightning bolt rendered as styled text -->
                        <span style="font-size:1.6rem;font-weight:900;color:#F7FE5E;
                                     font-family:Arial,sans-serif;line-height:46px;
                                     display:block;letter-spacing:-2px;">&#9889;</span>
                      </td>
                    </tr>
                  </table>
                </td>
                <!-- Brand text column -->
                <td style="vertical-align:middle;text-align:left;">
                  <!-- SFS label -->
                  <div style="font-size:0.6rem;font-weight:800;letter-spacing:3px;
                               text-transform:uppercase;color:rgba(247,254,94,0.85);
                               font-family:Arial,sans-serif;line-height:1;margin-bottom:3px;">
                    SFS College
                  </div>
                  <!-- Blixtro wordmark -->
                  <div style="font-size:1.45rem;font-weight:900;letter-spacing:0.5px;
                               color:#ffffff;font-family:Arial,sans-serif;line-height:1;">
                    Blixtro
                  </div>
                </td>
              </tr>
            </table>

            <div style="color:rgba(255,255,255,0.55);font-size:0.75rem;letter-spacing:1px;
                        text-transform:uppercase;margin-top:4px;font-family:Arial,sans-serif;">
              Inventory &bull; Booking &bull; Reporting
            </div>
          </td>
        </tr>

        <!-- Title band -->
        <tr>
          <td style="background:linear-gradient(90deg,#0ea5e9,#0284c7);padding:14px 36px;">
            <div style="color:#fff;font-size:0.72rem;font-weight:700;letter-spacing:1.2px;
                        text-transform:uppercase;margin-bottom:2px;">
              Forwarded Requirements
            </div>
            <div style="color:#fff;font-size:1.05rem;font-weight:700;">{_esc(subject)}</div>
          </td>
        </tr>

        <!-- Greeting -->
        <tr>
          <td style="padding:28px 36px 0;">
            <p style="margin:0 0 18px;font-size:0.95rem;color:#1e293b;line-height:1.6;">
              Dear <strong>{_esc(department)}</strong> Team,
            </p>
            <p style="margin:0 0 18px;font-size:0.9rem;color:#475569;line-height:1.6;">
              Please find below the room booking requirements forwarded from the
              <strong>AURA Command System</strong>. Kindly review and action as required.
            </p>
          </td>
        </tr>

        <!-- Booking info card -->
        <tr>
          <td style="padding:0 36px 20px;">
            <table width="100%" cellpadding="0" cellspacing="0"
                   style="background:#f8fafc;border:1px solid #e2e8f0;
                           border-radius:12px;overflow:hidden;font-size:0.85rem;">
              <tr style="background:#f1f5f9;">
                <td colspan="2" style="padding:10px 16px;font-weight:700;font-size:0.75rem;
                                        color:#64748b;text-transform:uppercase;letter-spacing:0.5px;
                                        border-bottom:1px solid #e2e8f0;">
                  Booking Details
                </td>
              </tr>
              <tr>
                <td style="padding:9px 16px;color:#64748b;font-weight:600;
                            border-bottom:1px solid #f1f5f9;width:38%;">Room</td>
                <td style="padding:9px 16px;color:#1e293b;font-weight:700;
                            border-bottom:1px solid #f1f5f9;">{_esc(room_name)}</td>
              </tr>
              <tr>
                <td style="padding:9px 16px;color:#64748b;font-weight:600;
                            border-bottom:1px solid #f1f5f9;">Faculty</td>
                <td style="padding:9px 16px;color:#1e293b;
                            border-bottom:1px solid #f1f5f9;">{_esc(faculty_name)}</td>
              </tr>
              <tr>
                <td style="padding:9px 16px;color:#64748b;font-weight:600;
                            border-bottom:1px solid #f1f5f9;">Faculty Email</td>
                <td style="padding:9px 16px;color:#1e293b;
                            border-bottom:1px solid #f1f5f9;">{_esc(faculty_email)}</td>
              </tr>
              <tr>
                <td style="padding:9px 16px;color:#64748b;font-weight:600;
                            border-bottom:1px solid #f1f5f9;">Schedule</td>
                <td style="padding:9px 16px;color:#1e293b;
                            border-bottom:1px solid #f1f5f9;">{_esc(start_local)} – {_esc(end_local)}</td>
              </tr>
              <tr>
                <td style="padding:9px 16px;color:#64748b;font-weight:600;">Purpose</td>
                <td style="padding:9px 16px;color:#1e293b;">{_esc(purpose)}</td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- Requirements content -->
        <tr>
          <td style="padding:0 36px 28px;">
            <div style="background:#fafafa;border:1px solid #e2e8f0;border-radius:12px;
                         border-left:4px solid #0ea5e9;padding:20px 22px;">
              <div style="font-weight:700;font-size:0.78rem;color:#0284c7;text-transform:uppercase;
                           letter-spacing:0.5px;margin-bottom:14px;">
                Requirements / Specifications
              </div>
              {content_html}
            </div>
          </td>
        </tr>

        <!-- Action note -->
        <tr>
          <td style="padding:0 36px 28px;">
            <div style="background:#fffbeb;border:1px solid #fde68a;border-radius:10px;
                         padding:12px 16px;font-size:0.82rem;color:#92400e;">
              <strong>&#9888; Action Required:</strong>
              Please acknowledge receipt and confirm arrangements for the above requirements
              at your earliest convenience.
            </div>
          </td>
        </tr>

        <!-- Divider -->
        <tr>
          <td style="padding:0 36px;">
            <hr style="border:none;border-top:1px solid #f1f5f9;margin:0 0 20px;">
          </td>
        </tr>

        <!-- Sender info -->
        <tr>
          <td style="padding:0 36px 28px;font-size:0.82rem;color:#64748b;line-height:1.7;">
            This email was forwarded by <strong style="color:#1e293b;">{_esc(sender_name)}</strong>
            ({_esc(sender_role)}) via the Blixtro AURA Command System.<br>
            This is an automated message — please do not reply to this email directly.
          </td>
        </tr>

        <!-- Footer -->
        <tr>
          <td style="background:#f8fafc;border-top:1px solid #f1f5f9;
                      padding:18px 36px;text-align:center;">
            <div style="font-size:0.75rem;color:#94a3b8;line-height:1.6;">
              <strong style="color:#475569;">Blixtro</strong> — SFS College, Autonomous Inventory &amp;
              Booking System<br>
              &copy; {_tz.now().year} Blixtro - SFS College, Autonomous. All rights reserved.
            </div>
          </td>
        </tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""

    # Plain-text fallback
    plain_lines = [
        f'Room Booking Requirements — {room_name}',
        f'Schedule : {start_local} – {end_local}',
        f'Faculty  : {faculty_name} <{faculty_email}>',
        f'Purpose  : {purpose}',
        '',
        'REQUIREMENTS',
        '─' * 40,
    ]
    for blk in blocks:
        if blk.get('type') == 'paragraph':
            plain_lines.append(blk.get('text', ''))
        elif blk.get('type') == 'table':
            for row in blk.get('rows', []):
                plain_lines.append(' | '.join(str(c) for c in row))
            plain_lines.append('')
    plain_lines += [
        '',
        '─' * 40,
        f'Forwarded by {sender_name} ({sender_role}) via Blixtro SFS College, Autonomous.',
        'This is an automated message — please do not reply directly.',
    ]
    plain_body = '\n'.join(plain_lines)

    # ── Send email ────────────────────────────────────────────────────────────
    try:
        from_email = getattr(_s, 'DEFAULT_FROM_EMAIL', 'noreply@sfscollege.in')
        msg = EmailMultiAlternatives(
            subject=subject,
            body=plain_body,
            from_email=from_email,
            to=[email],
        )
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=False)   
    except Exception as e:
        import traceback
        print(f'[forward_booking_requirements] Email send failed: {e}\n{traceback.format_exc()}')
        return JsonResponse(
            {'error': f'Email could not be sent: {str(e)}. Check your email settings.'},
            status=500
        )

    return JsonResponse({'status': 'success', 'message': f'Requirements forwarded to {email}.'})
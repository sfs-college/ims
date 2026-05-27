from django.forms import ValidationError
from django.shortcuts import redirect, get_object_or_404, render, reverse
from django.urls import reverse_lazy
from django.views.generic import ListView, UpdateView, DeleteView, TemplateView, CreateView, View
from inventory.models import Category, Vendor, Purchase, Room, Brand, Item, System, SystemComponent, Issue, ItemGroup, ItemGroupItem, RoomSettings, StockRequest, Archive, IssueTimeExtensionRequest, ItemConfiguration
from inventory.forms.room_incharge import CategoryForm, BrandForm, ItemForm, ItemPurchaseForm, PurchaseForm, PurchaseUpdateForm, SystemForm, SystemComponentForm, ItemGroupForm, ItemGroupItemForm, RoomSettingsForm, StockRequestForm 
from django.contrib import messages
from django.views.generic.edit import FormView
from inventory.forms.room_incharge import SystemComponentArchiveForm, ItemArchiveForm, RoomUpdateForm, IssueTimeExtensionForm
from inventory.forms.room_incharge import PurchaseCompleteForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, HttpResponseForbidden
from django.template.loader import render_to_string
import pandas as pd
import io
import logging
from datetime import timedelta
from django.utils import timezone
from datetime import datetime, date
from openpyxl.utils import datetime as xl_datetime
from decimal import Decimal, InvalidOperation
from django.forms.models import model_to_dict
from django.db import connection
from django.views.generic import FormView, View
from django.urls import reverse
from django.http import JsonResponse
from django.core.exceptions import PermissionDenied
from django.db.models import Q, F, Count

logger = logging.getLogger(__name__)


class CategoryListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/category_list.html'
    model = Category
    context_object_name = 'categories'

    def get_queryset(self):
        room_slug = self.kwargs['room_slug']
        profile = self.request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        return super().get_queryset().filter(room=room, organisation=profile.org).annotate(
            item_count=Count('item')
        ).order_by('category_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room_slug = self.kwargs['room_slug']
        profile = self.request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        context['room_slug'] = room_slug
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    def dispatch(self, request, *args, **kwargs):
        profile = getattr(request.user, 'profile', None)
        if not profile or not (profile.is_central_admin or profile.is_sub_admin):
            return HttpResponseForbidden("Only administrators can add or edit categories.")
        return super().dispatch(request, *args, **kwargs)

    model = Category
    template_name = 'room_incharge/category_update.html'
    form_class = CategoryForm
    success_url = reverse_lazy('room_incharge:category_list')
    slug_field = 'slug'
    slug_url_kwarg = 'category_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:category_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        category = form.save(commit=False)
        profile = self.request.user.profile
        category.organisation = profile.org
        category.room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        category.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.request.user.profile
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class CategoryCreateView(LoginRequiredMixin, CreateView):
    def dispatch(self, request, *args, **kwargs):
        profile = getattr(request.user, 'profile', None)
        if not profile or not (profile.is_central_admin or profile.is_sub_admin):
            return HttpResponseForbidden("Only administrators can add or edit categories.")
        return super().dispatch(request, *args, **kwargs)

    model = Category
    template_name = 'room_incharge/category_create.html'
    form_class = CategoryForm
    success_url = reverse_lazy('room_incharge:category_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:category_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        category = form.save(commit=False)
        profile = self.request.user.profile
        category.organisation = profile.org
        category.room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        category.save()
        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        profile = self.request.user.profile
        kwargs['initial']['room'] = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.request.user.profile
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class RoomDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'room_incharge/room_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room_slug = self.kwargs['room_slug']
        room = get_object_or_404(Room, slug=room_slug, incharge=self.request.user.profile)
        context['room'] = room
        context['room_slug'] = room_slug
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class RoomUpdateView(LoginRequiredMixin, UpdateView):
    model = Room
    template_name = 'room_incharge/room_update.html'
    form_class = RoomUpdateForm
    slug_field = 'slug'
    slug_url_kwarg = 'room_slug'

    def get_object(self, queryset=None):
        return get_object_or_404(Room, slug=self.kwargs['room_slug'], incharge=self.request.user.profile)

    def get_success_url(self):
        return reverse_lazy('room_incharge:room_dashboard', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        room = form.save(commit=False)
        room.organisation = self.request.user.profile.org
        room.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = self.get_object()
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class BrandListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/brand_list.html'
    model = Brand
    context_object_name = 'brands'

    def get_queryset(self):
        room_slug = self.kwargs['room_slug']
        profile = self.request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        return super().get_queryset().filter(room=room, organisation=profile.org).annotate(
            item_count=Count('item')
        ).order_by('brand_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room_slug = self.kwargs['room_slug']
        profile = self.request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        context['room_slug'] = room_slug
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class BrandCreateView(LoginRequiredMixin, CreateView):
    def dispatch(self, request, *args, **kwargs):
        profile = getattr(request.user, 'profile', None)
        if not profile or not (profile.is_central_admin or profile.is_sub_admin):
            return HttpResponseForbidden("Only administrators can add or edit brands.")
        return super().dispatch(request, *args, **kwargs)

    model = Brand
    template_name = 'room_incharge/brand_create.html'
    form_class = BrandForm
    success_url = reverse_lazy('room_incharge:brand_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:brand_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        brand = form.save(commit=False)
        profile = self.request.user.profile
        brand.organisation = profile.org
        brand.room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        brand.save()
        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        profile = self.request.user.profile
        kwargs['initial']['room'] = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.request.user.profile
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class BrandUpdateView(LoginRequiredMixin, UpdateView):
    def dispatch(self, request, *args, **kwargs):
        profile = getattr(request.user, 'profile', None)
        if not profile or not (profile.is_central_admin or profile.is_sub_admin):
            return HttpResponseForbidden("Only administrators can add or edit brands.")
        return super().dispatch(request, *args, **kwargs)

    model = Brand
    template_name = 'room_incharge/brand_update.html'
    form_class = BrandForm
    success_url = reverse_lazy('room_incharge:brand_list')
    slug_field = 'slug'
    slug_url_kwarg = 'brand_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:brand_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        brand = form.save(commit=False)
        profile = self.request.user.profile
        brand.organisation = profile.org
        brand.room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        brand.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.request.user.profile
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class ItemListView(LoginRequiredMixin, ListView):
    model = Item
    template_name = 'room_incharge/item_list.html'
    context_object_name = 'items'

    def get_queryset(self):
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        return Item.objects.filter(room=room).order_by('item_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        context['room'] = room
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]

        table_names = set(connection.introspection.table_names())
        pending_items = set()
        if "inventory_stockrequest" in table_names:
            pending_items = set(
                StockRequest.objects
                .filter(room=room, status="pending")
                .values_list("item_id", flat=True)
            )
        context["pending_items"] = pending_items
        profile = getattr(self.request.user, 'profile', None)
        context['is_central_admin'] = bool(profile and profile.is_central_admin and not profile.is_sub_admin)
        context['is_sub_admin'] = bool(profile and profile.is_sub_admin)
        return context


class ItemCreateView(LoginRequiredMixin, CreateView):
    model = Item
    template_name = 'room_incharge/item_create.html'
    form_class = ItemForm
    success_url = reverse_lazy('room_incharge:item_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:item_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        form.fields['category'].queryset = Category.objects.filter(room=room)
        form.fields['brand'].queryset = Brand.objects.filter(room=room)
        return form

    def form_valid(self, form):
        obj = form.save(commit=False)
        profile = getattr(self.request.user, "profile", None)

        if profile:
            obj.organisation = profile.org
            obj.created_by = profile

        obj.room = Room.objects.get(slug=self.kwargs['room_slug'])
        obj.save()

        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial']['room'] = Room.objects.get(slug=self.kwargs['room_slug'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SubmitStockRequestView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        room = get_object_or_404(Room, slug=kwargs["room_slug"])
        item_id         = request.POST.get("item_id", "").strip()
        requested_count = request.POST.get("requested_count", "").strip()
        reason          = request.POST.get("reason", "").strip()

        if not item_id or not requested_count or not reason:
            return JsonResponse(
                {"status": "error", "error": "All fields are required."},
                status=400,
            )

        try:
            requested_count = int(requested_count)
            if requested_count < 1:
                raise ValueError
        except (ValueError, TypeError):
            return JsonResponse(
                {"status": "error", "error": "Unit count must be a positive number."},
                status=400,
            )

        item = get_object_or_404(Item, id=item_id, room=room)

        if StockRequest.objects.filter(item=item, status="pending").exists():
            return JsonResponse(
                {"status": "error", "error": "A stock request for this item is already pending approval."},
                status=400,
            )

        profile = getattr(self.request.user, "profile", None)
        if not profile:
            return JsonResponse(
                {"status": "error", "error": "User profile not found."},
                status=403,
            )

        StockRequest.objects.create(
            item            = item,
            room            = room,
            requested_by    = profile,
            requested_count = requested_count,
            reason          = reason,
            status          = "pending",
        )

        return JsonResponse({"status": "success"})


class ItemArchiveView(LoginRequiredMixin, FormView):
    template_name = 'room_incharge/item_archive.html'
    form_class = ItemArchiveForm

    def get_success_url(self):
        return reverse_lazy(
            'room_incharge:item_list',
            kwargs={'room_slug': self.kwargs['room_slug']}
        )

    def form_valid(self, form):
        item = get_object_or_404(Item, slug=self.kwargs['item_slug'])
        count = form.cleaned_data['count']

        if count <= 0:
            form.add_error("count", "Count must be positive.")
            return self.form_invalid(form)

        if item.available_count < count:
            form.add_error("count", "Available count is lower than the number to archive.")
            return self.form_invalid(form)

        category = form.cleaned_data.get("archive_category", "serviceable")

        Archive.objects.create(
            organisation=item.organisation,
            department=item.department,
            room=item.room,
            item=item,
            count=count,
            archive_type='consumption',
            archive_category=category,
            archive_status='archived',
            remark=form.cleaned_data.get("remark", "")
        )

        item.archived_count += count
        if category == 'serviceable':
            item.serviceable_count += count
        else:
            item.unserviceable_count += count
        item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
        item.save(update_fields=["available_count", "archived_count", "serviceable_count", "unserviceable_count", "updated_on"])

        messages.success(self.request, "Item archived successfully.")
        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial']['item_slug'] = self.kwargs['item_slug']
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/system_list.html'
    model = Item
    context_object_name = 'items'

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    def get_room(self):
        profile = self.request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            return get_object_or_404(Room, slug=self.kwargs['room_slug'], organisation=profile.org)
        return get_object_or_404(Room, slug=self.kwargs['room_slug'], incharge=profile)

    def get_queryset(self):
        room = self.get_room()
        return Item.objects.filter(room=room).select_related('category', 'brand').order_by('item_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = self.get_room()
        items = context['items']
        context['room_slug'] = self.kwargs['room_slug']
        context['room'] = room
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]

        # Build active/inactive entries from item fields minus counts already assigned to systems
        from django.db.models import Count
        assigned_counts = (
            SystemComponent.objects
            .filter(component_item__room=room)
            .values('component_item_id', 'status')
            .annotate(count=Count('id'))
        )
        
        assigned_map = {}
        for entry in assigned_counts:
            item_id = entry['component_item_id']
            status = entry['status']
            count = entry['count']
            if item_id not in assigned_map:
                assigned_map[item_id] = {'active': 0, 'inactive': 0, 'under_maintenance': 0}
            
            key = 'under_maintenance' if status in ('under_maintenance', 'disposed') else status
            if key in assigned_map[item_id]:
                assigned_map[item_id][key] += count

        active_items = []
        inactive_items = []
        archived_items = []

        for item in items:
            assigned = assigned_map.get(item.id, {'active': 0, 'inactive': 0, 'under_maintenance': 0})
            is_assigned_to_system = (assigned['active'] > 0 or assigned['inactive'] > 0 or assigned['under_maintenance'] > 0)
            
            if is_assigned_to_system:
                loose_active = 0
                loose_inactive = 0
                loose_archived = 0
            else:
                loose_active = max(0, item.active_count)
                loose_inactive = max(0, item.inactive_count)
                loose_archived = max(0, item.archived_count)
            
            if loose_active > 0:
                active_items.append({'item': item, 'count': loose_active})
            if loose_inactive > 0:
                inactive_items.append({'item': item, 'count': loose_inactive})
            if loose_archived > 0:
                archived_items.append({'item': item, 'count': loose_archived})

        systems = (
            System.objects
            .filter(room=room)
            .annotate(
                active_component_count=Count('systemcomponent', filter=Q(systemcomponent__status='active')),
                inactive_component_count=Count('systemcomponent', filter=Q(systemcomponent__status='inactive')),
                archived_component_count=Count('systemcomponent', filter=Q(systemcomponent__status__in=['under_maintenance', 'disposed'])),
            )
            .prefetch_related('systemcomponent_set__component_item')
            .order_by('-created_on', 'system_name')
        )

        context['active_items'] = active_items
        context['inactive_items'] = inactive_items
        context['archived_items'] = archived_items
        context['systems'] = systems
        return context

class SystemConfigurationView(LoginRequiredMixin, View):
    def post(self, request, room_slug):
        import json
        from inventory.models import SystemConfiguration
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        system_slugs = data.get('system_slugs', [])
        configuration = data.get('configuration', '').strip()

        if not system_slugs:
            return JsonResponse({'error': 'No systems selected'}, status=400)
        if not configuration:
            return JsonResponse({'error': 'Configuration text is required'}, status=400)

        room = get_object_or_404(Room, slug=room_slug)
        updated = []
        for slug in system_slugs:
            system = System.objects.filter(slug=slug, room=room).first()
            if system:
                SystemConfiguration.objects.update_or_create(
                    system=system,
                    defaults={'configuration': configuration}
                )
                updated.append(system.system_name)

        return JsonResponse({'success': True, 'updated': updated})


class SystemConfigurationDetailView(LoginRequiredMixin, View):
    def get(self, request, room_slug, system_slug):
        from inventory.models import SystemConfiguration
        system = get_object_or_404(System, slug=system_slug, room__slug=room_slug)
        try:
            config = system.configuration
            return JsonResponse({'configuration': config.configuration, 'updated_on': config.updated_on.strftime('%d %b %Y, %H:%M')})
        except SystemConfiguration.DoesNotExist:
            return JsonResponse({'configuration': None})

class SystemCreateView(LoginRequiredMixin, CreateView):
    model = System
    template_name = 'room_incharge/system_create.html'
    form_class = SystemForm
    success_url = reverse_lazy('room_incharge:system_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:system_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        system = form.save(commit=False)
        system.organisation = self.request.user.profile.org
        system.room = Room.objects.get(slug=self.kwargs['room_slug'])
        system.department = system.room.department
        system.save()
        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial']['room'] = Room.objects.get(slug=self.kwargs['room_slug'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemUpdateView(LoginRequiredMixin, UpdateView):
    model = System
    template_name = 'room_incharge/system_update.html'
    form_class = SystemForm
    success_url = reverse_lazy('room_incharge:system_list')
    slug_field = 'slug'
    slug_url_kwarg = 'system_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:system_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        system = form.save(commit=False)
        system.organisation = self.request.user.profile.org
        system.room = Room.objects.get(slug=self.kwargs['room_slug'])
        system.department = system.room.department
        system.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemComponentListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/system_component_list.html'
    model = SystemComponent
    context_object_name = 'components'

    def get_queryset(self):
        system_slug = self.kwargs['system_slug']
        return super().get_queryset().filter(system__slug=system_slug, system__organisation=self.request.user.profile.org)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['system_slug'] = self.kwargs['system_slug']
        context['room_slug'] = self.kwargs['room_slug']
        context['system'] = get_object_or_404(System, slug=self.kwargs['system_slug'])
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemComponentCreateView(LoginRequiredMixin, CreateView):
    model = SystemComponent
    template_name = 'room_incharge/system_component_create.html'
    form_class = SystemComponentForm
    success_url = reverse_lazy('room_incharge:system_component_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:system_component_list', kwargs={'room_slug': self.kwargs['room_slug'], 'system_slug': self.kwargs['system_slug']})

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        form.fields['component_item'].queryset = Item.objects.filter(room=room)
        return form

    def form_valid(self, form):
        component = form.save(commit=False)
        component.system = System.objects.get(slug=self.kwargs['system_slug'])
        try:
            component.save()
        except ValueError as e:
            form.add_error(None, str(e))
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        item = component.component_item
        Item.objects.filter(pk=item.pk).update(
            active_count=F('active_count') + 1,
            in_use=F('active_count') + 1,
            available_count=F('available_count') - 1,
        )

        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial']['system'] = System.objects.get(slug=self.kwargs['system_slug'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['system_slug'] = self.kwargs['system_slug']
        context['room_slug'] = self.kwargs['room_slug']
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemComponentUpdateView(LoginRequiredMixin, UpdateView):
    model = SystemComponent
    template_name = 'room_incharge/system_component_update.html'
    form_class = SystemComponentForm
    success_url = reverse_lazy('room_incharge:system_component_list')
    slug_field = 'slug'
    slug_url_kwarg = 'component_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:system_component_list', kwargs={'room_slug': self.kwargs['room_slug'], 'system_slug': self.kwargs['system_slug']})

    def form_valid(self, form):
        component = form.save(commit=False)
        component.system = System.objects.get(slug=self.kwargs['system_slug'])

        old_component = SystemComponent.objects.get(pk=component.pk)
        old_item = old_component.component_item

        try:
            component.save()
        except ValueError as e:
            form.add_error(None, str(e))
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        new_item = component.component_item
        if old_item != new_item:
            Item.objects.filter(pk=old_item.pk).update(
                active_count=F('active_count') - 1,
                in_use=F('active_count') - 1,
                available_count=F('available_count') + 1,
            )
            Item.objects.filter(pk=new_item.pk).update(
                active_count=F('active_count') + 1,
                in_use=F('active_count') + 1,
                available_count=F('available_count') - 1,
            )

        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['system_slug'] = self.kwargs['system_slug']
        context['room_slug'] = self.kwargs['room_slug']
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemComponentArchiveView(LoginRequiredMixin, FormView):
    template_name = 'room_incharge/system_component_archive.html'
    form_class = SystemComponentArchiveForm

    def get_success_url(self):
        return reverse_lazy(
            'room_incharge:system_component_list',
            kwargs={'room_slug': self.kwargs['room_slug'], 'system_slug': self.kwargs['system_slug']}
        )

    def form_valid(self, form):
        component = get_object_or_404(SystemComponent, slug=self.kwargs["component_slug"])
        item = component.component_item

        Archive.objects.create(
            organisation=item.organisation,
            department=item.department,
            room=item.room,
            item=item,
            count=1,
            archive_type='consumption',
            archive_category='serviceable',
            archive_status='archived',
            remark=form.cleaned_data.get("remark", "")
        )

        # Component was active (in_use), so decrement active_count
        item.active_count = max(0, item.active_count - 1)
        item.in_use = item.active_count
        item.archived_count += 1
        item.serviceable_count += 1
        item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
        item.save(update_fields=["active_count", "in_use", "archived_count", "serviceable_count", "available_count", "updated_on"])

        component.delete()

        messages.success(self.request, "Component archived successfully.")
        return redirect(self.get_success_url())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial']['component_slug'] = self.kwargs['component_slug']
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['system_slug'] = self.kwargs['system_slug']
        context['room_slug'] = self.kwargs['room_slug']
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class SystemImportView(LoginRequiredMixin, View):
    template_name = 'room_incharge/system_import.html'

    def get(self, request, *args, **kwargs):
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        room_settings = RoomSettings.objects.get_or_create(room=room)[0]
        return render(request, self.template_name, {
            'room_slug': self.kwargs['room_slug'],
            'room': room,
            'room_settings': room_settings,
        })

    def post(self, request, *args, **kwargs):
        import openpyxl
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        room_settings = RoomSettings.objects.get_or_create(room=room)[0]
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return render(request, self.template_name, {
                'room_slug': self.kwargs['room_slug'],
                'room': room,
                'room_settings': room_settings,
                'error': 'Please upload an Excel file.',
            })

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception:
            return render(request, self.template_name, {
                'room_slug': self.kwargs['room_slug'],
                'room': room,
                'room_settings': room_settings,
                'error': 'Invalid Excel file. Please use the provided template.',
            })

        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        required = ['System Name', 'Component Type', 'Item Name', 'Serial Number']
        missing = [h for h in required if h not in headers]
        if missing:
            return render(request, self.template_name, {
                'room_slug': self.kwargs['room_slug'],
                'room': room,
                'room_settings': room_settings,
                'error': f'Missing columns: {", ".join(missing)}. Please use the provided template.',
            })

        col = {name: idx for idx, name in enumerate(headers)}

        # Valid component types
        valid_types = [c[0] for c in SystemComponent.COMPONENT_TYPES]
        valid_statuses = [c[0] for c in SystemComponent.STATUS_CHOICES]

        # All items in this room for matching
        room_items = {item.item_name.strip().lower(): item for item in Item.objects.filter(room=room)}

        preview_systems = {}   # system_name -> {components: [], error: None}
        row_errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            def get_col(name):
                idx = col.get(name)
                if idx is None:
                    return ''
                val = row[idx]
                return str(val).strip() if val is not None else ''

            system_name = get_col('System Name')
            component_type = get_col('Component Type').lower()
            item_name = get_col('Item Name')
            serial_number = get_col('Serial Number')
            comp_status = get_col('Component Status') or 'active'

            if not system_name:
                row_errors.append({'row': row_num, 'issue': 'System Name is empty'})
                continue

            errors = []
            if component_type not in valid_types:
                errors.append(f'Invalid component type "{component_type}"')
            matched_item = room_items.get(item_name.strip().lower())
            if not matched_item:
                errors.append(f'Item "{item_name}" not found in this room')
            if comp_status not in valid_statuses:
                comp_status = 'active'

            if system_name not in preview_systems:
                preview_systems[system_name] = {'components': [], 'has_error': False}

            preview_systems[system_name]['components'].append({
                'component_type': component_type,
                'item_name': item_name,
                'item_id': matched_item.id if matched_item else None,
                'serial_number': serial_number,
                'status': comp_status,
                'errors': errors,
                'row': row_num,
            })
            if errors:
                preview_systems[system_name]['has_error'] = True

        # Store in session for confirm step
        import json
        session_data = {}
        for sname, sdata in preview_systems.items():
            session_data[sname] = {
                'components': [{
                    'component_type': c['component_type'],
                    'item_id': c['item_id'],
                    'serial_number': c['serial_number'],
                    'status': c['status'],
                } for c in sdata['components'] if not c['errors']],
            }
        request.session['system_import_data'] = json.dumps(session_data)
        request.session['system_import_room'] = str(room.slug)

        total_systems = len(preview_systems)
        total_components = sum(len(s['components']) for s in preview_systems.values())
        error_rows = sum(
            len([c for c in s['components'] if c['errors']])
            for s in preview_systems.values()
        )

        return render(request, self.template_name, {
            'room_slug': self.kwargs['room_slug'],
            'room': room,
            'room_settings': room_settings,
            'preview': preview_systems,
            'total_systems': total_systems,
            'total_components': total_components,
            'error_rows': error_rows,
            'show_preview': True,
        })


class SystemImportConfirmView(LoginRequiredMixin, View):

    def post(self, request, *args, **kwargs):
        import json
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])

        raw = request.session.get('system_import_data')
        session_room = request.session.get('system_import_room')

        if not raw or session_room != str(room.slug):
            messages.error(request, 'Import session expired. Please upload again.')
            return redirect(reverse_lazy('room_incharge:system_import', kwargs={'room_slug': room.slug}))

        data = json.loads(raw)
        created_systems = 0
        created_components = 0

        for system_name, sdata in data.items():
            if not sdata['components']:
                continue

            system = System.objects.create(
                organisation=room.organisation,
                department=room.department,
                room=room,
                system_name=system_name,
            )
            created_systems += 1

            for comp in sdata['components']:
                item = Item.objects.get(id=comp['item_id'])
                try:
                    SystemComponent.objects.create(
                        system=system,
                        component_item=item,
                        component_type=comp['component_type'],
                        serial_number=comp['serial_number'],
                        status=comp['status'],
                    )
                    Item.objects.filter(pk=item.pk).update(
                        active_count=F('active_count') + 1,
                        in_use=F('active_count') + 1,
                        available_count=F('available_count') - 1,
                    )
                    created_components += 1
                except Exception:
                    pass

        del request.session['system_import_data']
        del request.session['system_import_room']

        messages.success(request, f'Import complete: {created_systems} systems and {created_components} components created.')
        return redirect(reverse_lazy('room_incharge:system_list', kwargs={'room_slug': room.slug}))

class ArchiveListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/archive_list.html'
    model = Archive
    context_object_name = 'archives'

    def get_queryset(self):
        room_slug = self.kwargs['room_slug']
        # Active archives: not yet returned to stock (serviced = returned to available)
        return Archive.objects.filter(
            room__slug=room_slug,
            organisation=self.request.user.profile.org,
        ).exclude(
            archive_status='serviced'
        ).select_related('item').order_by('-archived_on')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        # History: only serviced (returned to available stock)
        context['history_archives'] = Archive.objects.filter(
            room=room,
            organisation=self.request.user.profile.org,
            archive_status='serviced',
        ).select_related('item').order_by('-archived_on')
        return context

class PurchaseListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/purchase_list.html'
    model = Purchase
    context_object_name = 'purchases'

    def get_queryset(self):
        room_slug = self.kwargs['room_slug']
        return super().get_queryset().filter(room__slug=room_slug, organisation=self.request.user.profile.org)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class PurchaseCreateView(LoginRequiredMixin, CreateView):
    model = Purchase
    template_name = 'room_incharge/purchase_create.html'
    form_class = PurchaseForm

    def get_success_url(self):
        return reverse_lazy('room_incharge:purchase_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        org = self.request.user.profile.org

        purchase = form.save(commit=False)
        purchase.organisation = org
        purchase.room = room
        purchase.status = "requested"

        if not purchase.vendor:
            form.add_error("vendor", "Vendor is required")
            return self.form_invalid(form)

        if not purchase.purchase_date:
            purchase.purchase_date = timezone.now().date()

        if purchase.item:
            item = purchase.item
            item.total_count += int(purchase.quantity)
            item.save(update_fields=["total_count", "updated_on"])

        purchase.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class PurchaseUpdateView(LoginRequiredMixin, UpdateView):
    model = Purchase
    template_name = 'room_incharge/purchase_update.html'
    form_class = PurchaseUpdateForm
    success_url = reverse_lazy('room_incharge:purchase_list')
    slug_field = 'slug'
    slug_url_kwarg = 'purchase_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:purchase_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        purchase = form.save(commit=False)
        if purchase.status != 'requested':
            purchase.status = 'requested'
        purchase.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class PurchaseNewItemCreateView(LoginRequiredMixin, CreateView):
    model = Purchase
    template_name = 'room_incharge/purchase_new_item_create.html'
    form_class = ItemPurchaseForm

    def get_success_url(self):
        return reverse_lazy('room_incharge:purchase_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        org = self.request.user.profile.org

        category = form.cleaned_data['category']
        brand = form.cleaned_data['brand']
        vendor = form.cleaned_data['vendor']
        qty = form.cleaned_data['quantity']
        unit = form.cleaned_data['unit_of_measure']

        item = Item.objects.create(
            organisation=org,
            department=room.department,
            room=room,
            category=category,
            brand=brand,
            item_name=form.cleaned_data['item_name'],
            item_description=form.cleaned_data.get('item_description', ''),
            serial_number=form.cleaned_data.get('serial_number', ''),
            purchase_model_code=form.cleaned_data.get('purchase_model_code', ''),
            vendor=vendor,
            total_count=int(qty),
            in_use=0,
            archived_count=0,
            is_listed=True
        )

        if not vendor:
            form.add_error("vendor", "Vendor is required")
            return self.form_invalid(form)

        purchase = Purchase.objects.create(
            organisation=org,
            room=room,
            item=item,
            vendor=vendor,
            quantity=qty,
            unit_of_measure=unit,
            cost=form.cleaned_data.get("cost"),
            cost_per_unit=form.cleaned_data.get("cost_per_unit"),
            invoice_number=form.cleaned_data.get("invoice_number"),
            purchase_date=form.cleaned_data.get("purchase_date") or timezone.now().date(),
            item_description=form.cleaned_data.get("item_description", item.item_name),
            remarks=form.cleaned_data.get("remarks", ""),
            status="requested",
        )

        if purchase.cost_per_unit and purchase.quantity:
            purchase.total_cost = purchase.cost_per_unit * Decimal(str(purchase.quantity))
            purchase.save(update_fields=['total_cost'])

        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context


class PurchaseCompleteView(LoginRequiredMixin, FormView):
    template_name = 'room_incharge/purchase_complete.html'
    form_class = PurchaseCompleteForm
    success_url = reverse_lazy('room_incharge:purchase_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:purchase_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        purchase = Purchase.objects.get(slug=self.kwargs['purchase_slug'])
        receipt = form.save(commit=False)
        receipt.purchase = purchase
        receipt.org = self.request.user.profile.org
        receipt.save()
        purchase.status = 'completed'
        purchase.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['purchase_slug'] = self.kwargs['purchase_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class PurchaseAddToStockView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        purchase = get_object_or_404(Purchase, slug=self.kwargs['purchase_slug'])
        if purchase.status == 'completed' and not purchase.added_to_stock:
            item = purchase.item
            item.total_count += purchase.quantity
            item.available_count += purchase.quantity
            if not item.is_listed:
                item.is_listed = True
            item.save()
            purchase.added_to_stock = True
            purchase.save()
            messages.success(request, f"Added {purchase.quantity} {purchase.unit_of_measure} to {item.item_name} stock.")
        return redirect('room_incharge:purchase_list', room_slug=self.kwargs['room_slug'])


class IssueListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/issue_list.html'
    model = Issue
    context_object_name = 'issues'

    def get_room(self):
        room = get_object_or_404(Room, slug=self.kwargs["room_slug"])
        profile = self.request.user.profile

        if (
            room.organisation == profile.org
            or room.incharge == profile
            or profile.is_sub_admin
            or profile.is_central_admin
        ):
            return room

        raise PermissionDenied("You do not have access to this room")

    def get_queryset(self):
        room = self.get_room()
        return Issue.objects.filter(room=room).order_by('-created_on')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = self.get_room()
        room_settings, _ = RoomSettings.objects.get_or_create(room=room)
        context.update({
            'room': room,
            'room_slug': room.slug,
            'room_settings': room_settings,
        })
        return context


class CentralIssuesListView(LoginRequiredMixin, ListView):
    """
    All issues across every room assigned to this incharge.
    No room_slug in the URL — the room is chosen via a ?room= dropdown.
    Query params:
      ?room=<slug>     – filter to a single room (only rooms with ≥1 open/in_progress issue are shown)
      ?q=<text>        – full-text search on subject / description / ticket_id
    Context:
      issues             – filtered Issue queryset
      rooms_with_counts  – list of {room, issue_count} for the dropdown
      selected_room      – Room object (if ?room= is set)
    """
    template_name = 'room_incharge/issue_list.html'
    model         = Issue
    context_object_name = 'issues'
    paginate_by   = 25

    def get_base_qs(self):
        profile = self.request.user.profile
        rooms   = profile.rooms_incharge.all()
        return Issue.objects.filter(room__in=rooms)

    def get_rooms_with_counts(self):
        """Return rooms the incharge manages that have at least one open/in_progress issue."""
        profile = self.request.user.profile
        return (
            Room.objects
            .filter(incharge=profile)
            .annotate(issue_count=Count('issue', filter=Q(issue__status__in=['open', 'in_progress'])))
            .filter(issue_count__gt=0)
            .order_by('room_name')
        )

    def get_queryset(self):
        qs   = self.get_base_qs()
        slug = self.request.GET.get('room', '').strip()
        q    = self.request.GET.get('q', '').strip()

        if slug:
            qs = qs.filter(room__slug=slug)

        if q:
            qs = qs.filter(
                Q(subject__icontains=q)
                | Q(description__icontains=q)
                | Q(ticket_id__icontains=q)
            )

        return qs.select_related('room', 'room__incharge', 'assigned_to').order_by('-created_on')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        slug    = self.request.GET.get('room', '').strip()
        search  = self.request.GET.get('q', '').strip()

        selected_room = None
        room_settings = RoomSettings.objects.none()
        if slug:
            try:
                selected_room = Room.objects.get(slug=slug, incharge=self.request.user.profile)
                room_settings = RoomSettings.objects.get_or_create(room=selected_room)[0]
            except Room.DoesNotExist:
                pass

        context.update({
            'room':               selected_room,
            'room_slug':          slug,
            'room_settings':      room_settings,
            'rooms_with_counts':  self.get_rooms_with_counts(),
            'selected_room_slug': slug,
            'search_query':       search,
            'is_central_view':    True,
        })
        return context


class MarkInProgressView(LoginRequiredMixin, View):
    def post(self, request, room_slug, pk):
        profile = request.user.profile
        room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        if room.incharge != profile:
            return HttpResponseForbidden("You are not allowed to update this issue.")

        issue = get_object_or_404(Issue, pk=pk, room=room)
        issue.status = "in_progress"
        issue.resolved = False
        issue.escalation_level = 0
        issue.updated_on = timezone.now()
        issue.save(update_fields=["status", "resolved", "escalation_level", "updated_on"])

        messages.success(request, f"Issue {issue.ticket_id} marked as IN PROGRESS.")
        return redirect("room_incharge:issue_list", room_slug=room.slug)


class MarkResolvedView(LoginRequiredMixin, View):
    def post(self, request, room_slug, pk):
        room = get_object_or_404(Room, slug=room_slug, organisation=request.user.profile.org)
        issue = get_object_or_404(Issue, pk=pk, room=room)
        issue.status = "closed"
        issue.resolved = True
        issue.save(update_fields=["status", "resolved", "updated_on"])
        return redirect('room_incharge:issue_list', room_slug=room.slug)


class MarkUnresolvedView(LoginRequiredMixin, View):
    def post(self, request, room_slug, pk):
        room = get_object_or_404(Room, slug=room_slug, organisation=request.user.profile.org)
        issue = get_object_or_404(Issue, pk=pk, room=room)
        issue.status = "open"
        issue.resolved = False
        issue.save(update_fields=["status", "resolved", "updated_on"])
        return redirect('room_incharge:issue_list', room_slug=room.slug)


class CloseIssueView(LoginRequiredMixin, View):
    def post(self, request, room_slug, pk):
        profile = request.user.profile
        room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)

        if not (room.incharge == profile or profile.is_sub_admin or profile.is_central_admin):
            return HttpResponseForbidden("Not allowed.")

        issue = get_object_or_404(Issue, pk=pk, room=room)
        closure_reason = request.POST.get('closure_reason', '').strip()

        if not closure_reason:
            messages.error(request, 'A closure reason is required.')
            return redirect('room_incharge:issue_list', room_slug=room.slug)

        issue.status = 'closed'
        issue.resolved = False
        issue.closure_reason = closure_reason
        issue.save(update_fields=['status', 'resolved', 'closure_reason', 'updated_on'])

        messages.success(request, f'Issue {issue.ticket_id} closed.')
        return redirect('room_incharge:issue_list', room_slug=room.slug)


class ItemGroupListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/item_group_list.html'
    model = ItemGroup
    context_object_name = 'item_groups'

    def get_queryset(self):
        room_slug = self.kwargs['room_slug']
        return super().get_queryset().filter(room__slug=room_slug, organisation=self.request.user.profile.org)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class ItemGroupCreateView(LoginRequiredMixin, CreateView):
    model = ItemGroup
    template_name = 'room_incharge/item_group_create.html'
    form_class = ItemGroupForm
    success_url = reverse_lazy('room_incharge:item_group_list')

    def get_success_url(self):
        return reverse_lazy('room_incharge:item_group_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        item_group = form.save(commit=False)
        item_group.organisation = self.request.user.profile.org
        item_group.room = Room.objects.get(slug=self.kwargs['room_slug'])
        item_group.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class ItemGroupItemCreateView(LoginRequiredMixin, CreateView):
    model = ItemGroupItem
    template_name = 'room_incharge/item_group_item_create.html'
    form_class = ItemGroupItemForm

    def get_success_url(self):
        return reverse_lazy('room_incharge:item_group_item_list', kwargs={'room_slug': self.kwargs['room_slug'], 'item_group_slug': self.kwargs['item_group_slug']})

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        form.fields['item'].queryset = Item.objects.filter(room=room)
        return form

    def form_valid(self, form):
        item_group_item = form.save(commit=False)
        item_group_item.item_group = ItemGroup.objects.get(slug=self.kwargs['item_group_slug'])
        item = item_group_item.item

        try:
            item_group_item.save()
            Item.objects.filter(pk=item.pk).update(
                available_count=F('available_count') - item_group_item.qty,
                in_use=F('in_use') + item_group_item.qty,
            )
        except ValueError as e:
            form.add_error(None, str(e))
            return self.form_invalid(form)

        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['room_slug'] = self.kwargs['room_slug']
        context['item_group_slug'] = self.kwargs['item_group_slug']
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class ItemGroupItemListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/item_group_item_list.html'
    model = ItemGroupItem
    context_object_name = 'item_group_items'

    def get_queryset(self):
        item_group_slug = self.kwargs['item_group_slug']
        return super().get_queryset().filter(item_group__slug=item_group_slug, item_group__organisation=self.request.user.profile.org)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['room_slug'] = self.kwargs['room_slug']
        context['item_group_slug'] = self.kwargs['item_group_slug']
        context['item_group'] = get_object_or_404(ItemGroup, slug=self.kwargs['item_group_slug'])
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class ItemGroupUpdateView(LoginRequiredMixin, UpdateView):
    model = ItemGroup
    template_name = 'room_incharge/item_group_update.html'
    form_class = ItemGroupForm
    success_url = reverse_lazy('room_incharge:item_group_list')
    slug_field = 'slug'
    slug_url_kwarg = 'item_group_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:item_group_list', kwargs={'room_slug': self.kwargs['room_slug']})

    def form_valid(self, form):
        item_group = form.save(commit=False)
        item_group.organisation = self.request.user.profile.org
        item_group.room = Room.objects.get(slug=self.kwargs['room_slug'])
        item_group.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class ItemGroupItemUpdateView(LoginRequiredMixin, UpdateView):
    model = ItemGroupItem
    template_name = 'room_incharge/item_group_item_update.html'
    form_class = ItemGroupItemForm
    slug_field = 'slug'
    slug_url_kwarg = 'item_group_item_slug'

    def get_success_url(self):
        return reverse_lazy('room_incharge:item_group_item_list', kwargs={'room_slug': self.kwargs['room_slug'], 'item_group_slug': self.kwargs['item_group_slug']})

    def form_valid(self, form):
        item_group_item = form.save(commit=False)
        item = item_group_item.item
        old_qty = ItemGroupItem.objects.get(pk=item_group_item.pk).qty

        item.available_count += old_qty - item_group_item.qty
        item.in_use -= old_qty - item_group_item.qty
        item.save()

        item_group_item.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['room_slug'] = self.kwargs['room_slug']
        context['item_group_slug'] = self.kwargs['item_group_slug']
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context


class RoomSettingsView(LoginRequiredMixin, UpdateView):
    model = RoomSettings
    template_name = 'room_incharge/room_settings.html'
    form_class = RoomSettingsForm
    success_url = reverse_lazy('room_incharge:room_dashboard')

    def get_object(self):
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        return RoomSettings.objects.get_or_create(room=room)[0]

    def get_success_url(self):
        return reverse_lazy('room_incharge:room_settings', kwargs={'room_slug': self.kwargs['room_slug']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['room_slug'] = self.kwargs['room_slug']
        room = Room.objects.get(slug=self.kwargs['room_slug'])
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        return context

class RoomReportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        room_slug = self.kwargs['room_slug']
        room = get_object_or_404(Room, slug=room_slug)
        room_settings = RoomSettings.objects.get_or_create(room=room)[0]
        format_type = request.GET.get('format', 'pdf')

        def fmt_datetime(dt):
            if not dt:
                return ''
            if isinstance(dt, date) and not isinstance(dt, datetime):
                return dt.strftime('%b. %d, %Y')
            try:
                dt_local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
            except Exception:
                dt_local = dt
            hour = dt_local.strftime('%I').lstrip('0') or '0'
            minute = dt_local.strftime('%M')
            ampm = dt_local.strftime('%p')
            ampm = 'a.m.' if ampm == 'AM' else 'p.m.'
            return f"{dt_local.strftime('%b. %d, %Y')}, {hour}:{minute} {ampm}"

        def safe_str(value):
            if value is None:
                return ''
            try:
                return str(value)
            except Exception:
                return ''

        systems_qs = System.objects.filter(room=room) if room_settings.systems_tab else None

        # Build system_configs as a list (safe for Django template iteration)
        # Each entry: {'name': str, 'rows': [{'spec': str, 'value': str}, ...]}
        system_configs_list = []
        if systems_qs is not None:
            from inventory.models import SystemConfiguration
            import json as _json
            for s in systems_qs:
                try:
                    cfg_text = s.configuration.configuration
                    try:
                        rows = _json.loads(cfg_text)
                        if not isinstance(rows, list):
                            raise ValueError
                    except Exception:
                        rows = [{'spec': 'Configuration', 'value': cfg_text}]
                except Exception:
                    rows = []
                if rows:
                    system_configs_list.append({'name': s.system_name, 'rows': rows})

        class ExistsList(list):
            def exists(self):
                return len(self) > 0

        items_qs = Item.objects.filter(room=room).select_related(
            'category', 'brand', 'created_by__user', 'updated_by__user'
        ) if room_settings.items_tab else None

        items_list = ExistsList()
        if items_qs is not None:
            for item in items_qs:
                item.created_by_email = item.created_by.user.email if (item.created_by and hasattr(item.created_by, 'user')) else '—'
                item.created_on_str = timezone.localtime(item.created_on).strftime('%d %b %Y, %I:%M %p') if item.created_on else '—'
                item.updated_by_email = '—'
                item.updated_on_str = '—'
                if item.updated_on and (item.updated_on - item.created_on).total_seconds() > 2:
                    item.updated_by_email = item.updated_by_emails or (item.updated_by.user.email if (item.updated_by and hasattr(item.updated_by, 'user')) else '—')
                    item.updated_on_str = timezone.localtime(item.updated_on).strftime('%d %b %Y, %I:%M %p')
                items_list.append(item)

        context = {
            'room': room,
            'room_settings': room_settings,
            'categories': Category.objects.filter(room=room) if room_settings.categories_tab else None,
            'brands': Brand.objects.filter(room=room) if room_settings.brands_tab else None,
            'items': items_list if room_settings.items_tab else None,
            'systems': systems_qs,
            'system_configs': system_configs_list,   # list of {name, rows}
            'item_groups': ItemGroup.objects.filter(room=room) if room_settings.item_groups_tab else None,
            'system_components': SystemComponent.objects.filter(system__room=room) if room_settings.systems_tab else None,
            'purchases': Purchase.objects.filter(room=room),
            'issues': Issue.objects.filter(room=room),
        }

        if format_type == 'excel':
            excel_buffer = io.BytesIO()

            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                summary_rows = [{
                    'Room Name': f"{room.label} - {room.room_name}" if room.label else room.room_name,
                    'Incharge': f"{room.incharge.first_name} {room.incharge.last_name}" if getattr(room, 'incharge', None) else '',
                    'Department': getattr(room.department, 'department_name', '') if getattr(room, 'department', None) else '',
                    'Created On': fmt_datetime(getattr(room, 'created_on', None)),
                    'Updated On': fmt_datetime(getattr(room, 'updated_on', None)),
                }]
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Summary', index=False)

                if context['categories'] is not None and context['categories'].exists():
                    rows = [{
                        'Category Name': c.category_name,
                        'Created On': fmt_datetime(c.created_on),
                        'Updated On': fmt_datetime(c.updated_on)
                    } for c in context['categories']]
                    pd.DataFrame(rows).to_excel(writer, sheet_name='Categories', index=False)

                if context['brands'] is not None and context['brands'].exists():
                    rows = [{
                        'Brand Name': b.brand_name,
                        'Created On': fmt_datetime(b.created_on),
                        'Updated On': fmt_datetime(b.updated_on)
                    } for b in context['brands']]
                    pd.DataFrame(rows).to_excel(writer, sheet_name='Brands', index=False)

                if context['items'] is not None and context['items'].exists():
                    item_rows = []
                    for i, item in enumerate(context['items'], start=1):
                        opening_stock = max(item.total_count - item.available_count, 0)
                        arrival = getattr(item, 'arrival_receipts', 0)
                        consumed = getattr(item, 'consumed_stock_qty', 0)
                        closing = item.available_count
                        total = opening_stock + arrival
                        item_rows.append({
                            'Sl No': i,
                            'Product Code': item.product_code or '—',
                            'Date of Entry': fmt_datetime(item.created_on),
                            'Item Description': item.item_description or item.item_name,
                            'Category': safe_str(getattr(item.category, 'category_name', '')),
                            'Opening Stock Qty': opening_stock,
                            'Arrival / Receipts': arrival,
                            'Total': total,
                            'Consumed Stock/Issues Qty': consumed,
                            'Closing / Balance Qty': closing,
                            'Unit of Measure': getattr(item, 'unit_of_measure', 'Units'),
                            'Remarks': getattr(item, 'remarks', ''),
                            'Active': item.active_count,
                            'Inactive': item.inactive_count,
                            'Under Maintenance': item.serviceable_count,
                            'Not Serviceable': item.unserviceable_count,
                            'Created By': item.created_by_email,
                            'Updated By': item.updated_by_email,
                            'Updated On': item.updated_on_str,
                        })
                    pd.DataFrame(item_rows).to_excel(writer, sheet_name='Items', index=False)

                if context['systems'] is not None and context['systems'].exists():
                    rows = [{
                        'System Name': s.system_name,
                        'Created On': fmt_datetime(s.created_on),
                        'Updated On': fmt_datetime(s.updated_on)
                    } for s in context['systems']]
                    pd.DataFrame(rows).to_excel(writer, sheet_name='Systems', index=False)

                # System Configurations sheet
                if context['system_configs']:
                    cfg_rows = []
                    for cfg in context['system_configs']:
                        for row in cfg['rows']:
                            cfg_rows.append({
                                'System Name': cfg['name'],
                                'Specification': row.get('spec', ''),
                                'Value / Details': row.get('value', ''),
                            })
                    if cfg_rows:
                        pd.DataFrame(cfg_rows).to_excel(writer, sheet_name='System Configurations', index=False)

                if context['system_components'] is not None and context['system_components'].exists():
                    rows = [{
                        'System': safe_str(c.system),
                        'Component Type': c.component_type,
                        'Component Item': safe_str(c.component_item),
                        'Serial Number': getattr(c, 'serial_number', ''),
                        'Created On': fmt_datetime(c.created_on),
                        'Updated On': fmt_datetime(c.updated_on),
                    } for c in context['system_components']]
                    pd.DataFrame(rows).to_excel(writer, sheet_name='System Components', index=False)

                if context['item_groups'] is not None and context['item_groups'].exists():
                    rows = [{
                        'Item Group Name': g.item_group_name,
                        'Created On': fmt_datetime(g.created_on),
                        'Updated On': fmt_datetime(g.updated_on)
                    } for g in context['item_groups']]
                    pd.DataFrame(rows).to_excel(writer, sheet_name='Item Groups', index=False)

                if context['purchases'] is not None and context['purchases'].exists():
                    purchase_rows = []
                    for i, p in enumerate(context['purchases'], start=1):
                        purchase_rows.append({
                            'Sl No': i,
                            'Date of Purchase/Entry': fmt_datetime(p.purchase_date) or fmt_datetime(p.date_of_entry),
                            'Item Description': safe_str(getattr(p.item, 'item_description', p.item_description)),
                            'Category': safe_str(getattr(p.item.category, 'category_name', '')) if getattr(p, 'item', None) else '',
                            'Purchase ID/Model Code': p.purchase_id or safe_str(getattr(p.item, 'purchase_model_code', '')),
                            'Serial No': safe_str(getattr(p.item, 'serial_number', '')),
                            'Quantity': p.quantity,
                            'Unit of Measure': p.unit_of_measure,
                            'Status': p.status,
                            'Vendor': safe_str(getattr(p.vendor, 'vendor_name', '')) if p.vendor else '',
                            'Remarks': safe_str(p.remarks),
                        })
                    pd.DataFrame(purchase_rows).to_excel(writer, sheet_name='Purchases', index=False)

                if context['issues'] is not None and context['issues'].exists():
                    rows = [{
                        'Subject': iss.subject,
                        'Description': iss.description,
                        'Resolved': 'Resolved' if iss.resolved else 'Unresolved',
                        'Created On': fmt_datetime(iss.created_on),
                        'Updated On': fmt_datetime(iss.updated_on)
                    } for iss in context['issues']]
                    pd.DataFrame(rows).to_excel(writer, sheet_name='Issues', index=False)

            # Autofit all columns in all sheets
            from openpyxl.utils import get_column_letter
            wb = writer.book
            for sheet in wb.worksheets:
                for col_cells in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col_cells[0].column)
                    for cell in col_cells:
                        try:
                            cell_len = len(str(cell.value)) if cell.value is not None else 0
                            if cell_len > max_len:
                                max_len = cell_len
                        except Exception:
                            pass
                    sheet.column_dimensions[col_letter].width = min(max_len + 4, 50)
                for row in sheet.iter_rows():
                    sheet.row_dimensions[row[0].row].height = 18

            excel_buffer.seek(0)
            # Sanitize filename: replace spaces and special chars for safe HTTP headers
            safe_name = room.room_name.replace(' ', '_').replace('/', '-')
            response = HttpResponse(
                excel_buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="{safe_name}_report.xlsx"; '
                f"filename*=UTF-8''{safe_name}_report.xlsx"
            )
            response['X-Content-Type-Options'] = 'nosniff'
            return response

        html_string = render_to_string('room_incharge/room_report.html', context)
        try:
            from weasyprint import HTML
            html = HTML(string=html_string)
            pdf = html.write_pdf()
        except Exception as e:
            logger.error(f"[RoomReportView] WeasyPrint PDF generation failed: {e}")
            return HttpResponse(
                f"PDF generation failed: {str(e)}. Please use Excel format instead.",
                status=500,
                content_type='text/plain'
            )

        safe_name = room.room_name.replace(' ', '_').replace('/', '-')
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_name}_report.pdf"; '
            f"filename*=UTF-8''{safe_name}_report.pdf"
        )
        response['X-Content-Type-Options'] = 'nosniff'
        return response

class IssueTimeExtensionRequestView(LoginRequiredMixin, View):
    def post(self, request, issue_id):
        issue = get_object_or_404(Issue, pk=issue_id)

        requested_extra_hours = request.POST.get("requested_extra_hours")
        reason = request.POST.get("reason")

        if not requested_extra_hours or not reason:
            messages.error(request, "Both additional time and reason are required.")
            return redirect("room_incharge:issue_list", room_slug=issue.room.slug)

        if issue.tat_deadline:
            remaining_seconds = (issue.tat_deadline - timezone.now()).total_seconds()
            current_tat_hours = max(int(remaining_seconds // 3600), 0)
        else:
            current_tat_hours = 48

        IssueTimeExtensionRequest.objects.create(
            issue=issue,
            requested_by=request.user.profile,
            current_tat_hours=current_tat_hours,
            requested_extra_hours=int(requested_extra_hours),
            reason=reason,
        )

        messages.success(request, "Time extension request submitted successfully.")
        return redirect("room_incharge:issue_list", room_slug=issue.room.slug)

class RoomInchargeNotificationsView(LoginRequiredMixin, View):
    template_name = "room_incharge/notifications.html"

    def get(self, request, *args, **kwargs):
        room_slug = kwargs["room_slug"]
        room = get_object_or_404(Room, slug=room_slug, incharge=request.user.profile)
        room_settings = RoomSettings.objects.get_or_create(room=room)[0]

        # Session key scoped to this room so dismissals don't bleed across rooms
        session_key = f"dismissed_notifs_{room_slug}"
        dismissed = request.session.get(session_key, {})

        def not_dismissed(prefix, obj_id):
            return str(obj_id) not in dismissed.get(prefix, [])

        stock_notifications = [
            r for r in (
                StockRequest.objects
                .filter(room=room, status__in=["approved", "rejected"])
                .select_related("item", "reviewed_by")
                .order_by("-created_on")[:50]
            )
            if not_dismissed("stock", r.id)
        ]

        try:
            assigned_notifications = [
                r for r in (
                    StockRequest.objects
                    .filter(room=room, status="approved", requested_by__isnull=True)
                    .select_related("item", "reviewed_by")
                    .order_by("-created_on")[:30]
                )
                if not_dismissed("assigned", r.id)
            ]
        except Exception:
            assigned_notifications = []

        issue_notifications = [
            r for r in (
                Issue.objects
                .filter(room=room)
                .order_by("-updated_on")[:50]
            )
            if not_dismissed("issue", r.id)
        ]

        # ── Cross-room summary (for the Issues section header / badge) ──
        profile      = request.user.profile
        managed_rooms = profile.rooms_incharge.all()
        all_open_issues_count = Issue.objects.filter(
            room__in=managed_rooms, status__in=['open', 'in_progress']
        ).count()
        rooms_with_counts_qs = (
            Room.objects
            .filter(incharge=profile)
            .annotate(_cnt=Count('issue', filter=Q(issue__status__in=['open', 'in_progress'])))
            .filter(_cnt__gt=0)
            .order_by('room_name')
        )
        rooms_with_counts = [
            {'room': r, 'issue_count': r.issue_count}
            for r in rooms_with_counts_qs
        ]

        context = {
            "room":                   room,
            "room_slug":              room_slug,
            "room_settings":          room_settings,
            "stock_notifications":    stock_notifications,
            "assigned_notifications": assigned_notifications,
            "issue_notifications":    issue_notifications,
            "all_open_issues_count":  all_open_issues_count,
            "rooms_with_counts":      rooms_with_counts,
            "managed_rooms":          managed_rooms,
            "central_issues_url":     reverse('room_incharge:central_issue_list'),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """
        AJAX endpoint to persist notification dismissals in session for room incharge.
        Body: { "action": "dismiss", "prefix": "stock", "id": 42 }
              { "action": "clear_all", "items": [{"prefix":"stock","id":42}, ...] }
        """
        import json
        room_slug = kwargs["room_slug"]
        room = get_object_or_404(Room, slug=room_slug, incharge=request.user.profile)

        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        session_key = f"dismissed_notifs_{room_slug}"
        dismissed = request.session.get(session_key, {})

        if data.get('action') == 'dismiss':
            prefix = data.get('prefix', '')
            obj_id = str(data.get('id', ''))
            if prefix and obj_id:
                if prefix not in dismissed:
                    dismissed[prefix] = []
                if obj_id not in dismissed[prefix]:
                    dismissed[prefix].append(obj_id)
                request.session[session_key] = dismissed
                request.session.modified = True
            return JsonResponse({'ok': True})

        elif data.get('action') == 'clear_all':
            # Persist each item passed from the frontend
            items = data.get('items', [])
            for item in items:
                prefix = item.get('prefix', '')
                obj_id = str(item.get('id', ''))
                if prefix and obj_id:
                    if prefix not in dismissed:
                        dismissed[prefix] = []
                    if obj_id not in dismissed[prefix]:
                        dismissed[prefix].append(obj_id)
            request.session[session_key] = dismissed
            request.session.modified = True
            return JsonResponse({'ok': True})

        return JsonResponse({'error': 'Unknown action'}, status=400)

def get_room_asset_tags(request, room_slug):
    from inventory.models import AssetTag
    from django.http import JsonResponse
    item_name = request.GET.get('item_name', '')
    tags = AssetTag.objects.filter(
        item_name=item_name,
        assigned_room__slug=room_slug,
    ).order_by('tag_id')
    return JsonResponse({'tags': [{'tag_id': t.tag_id} for t in tags]})


# ═══════════════════════════════════════════════════════════════
# ISSUE REMARK — Room incharge sends a progress update to reporter
# ═══════════════════════════════════════════════════════════════

class SendIssueRemarkView(LoginRequiredMixin, View):
    """
    Room incharge sends a progress remark to the issue reporter.
    - Does NOT change issue status or resolution state.
    - Saves the remark on the Issue model (incharge_remark field).
    - Emails the reporter with the update.
    """

    def post(self, request, room_slug, pk):
        profile = request.user.profile
        room    = get_object_or_404(Room, slug=room_slug, organisation=profile.org)

        if room.incharge != profile and not (profile.is_sub_admin or profile.is_central_admin):
            return HttpResponseForbidden("Not allowed.")

        issue  = get_object_or_404(Issue, pk=pk, room=room)
        remark = request.POST.get('incharge_remark', '').strip()

        if not remark:
            messages.error(request, 'Remark cannot be empty.')
            return redirect('room_incharge:issue_list', room_slug=room.slug)

        # Save remark on the issue
        issue.incharge_remark = remark
        issue.save(update_fields=['incharge_remark', 'updated_on'])

        # Email the reporter
        if issue.reporter_email:
            try:
                from inventory.email import safe_send_mail, build_email_shell
                incharge_name = f"{profile.first_name} {profile.last_name}".strip() or 'Room Incharge'
                safe_send_mail(
                    subject=f"[Blixtro] Update on Your Issue — {issue.ticket_id}",
                    message=(
                        f"Dear {issue.created_by or 'Student'},\n\n"
                        f"The room incharge has sent an update regarding your issue.\n\n"
                        f"Ticket ID : {issue.ticket_id}\n"
                        f"Subject   : {issue.subject}\n"
                        f"Status    : {issue.get_status_display()}\n\n"
                        f"Update from {incharge_name}:\n{remark}\n\n"
                        "Your issue is still being tracked. You will be notified once it is resolved.\n\n"
                        "Best regards,\nBlixtro — SFS College Inventory & Booking System"
                    ),
                    recipient_list=[issue.reporter_email],
                    html_message=build_email_shell(
                        title="Issue Progress Update",
                        intro_html=(
                            f"Dear <strong>{issue.created_by or 'Student'}</strong>, "
                            f"the room incharge has sent a progress update on your issue."
                        ),
                        sections=[
                            {
                                "title": "Issue Details",
                                "rows": [
                                    {"label": "Ticket ID", "value": issue.ticket_id},
                                    {"label": "Subject",   "value": issue.subject},
                                    {"label": "Status",    "value": issue.get_status_display()},
                                    {"label": "Room",      "value": issue.room.room_name},
                                ],
                            },
                            {
                                "title": f"Update from {incharge_name}",
                                "body_html": (
                                    f'<div style="white-space:pre-line;font-size:13px;line-height:1.7;color:#334155;">'
                                    f'{remark}'
                                    f'</div>'
                                ),
                            },
                        ],
                        outro_html="Your issue is still being tracked. You will be notified once it is fully resolved.",
                        accent="#6366f1",
                    ),
                )
            except Exception as _e:
                logger.error(f"[SendIssueRemark] Email failed: {_e}")

        messages.success(request, f'Remark sent to {issue.reporter_email}.')
        return redirect('room_incharge:issue_list', room_slug=room.slug)


# ═══════════════════════════════════════════════════════════════
# SYSTEMS KANBAN — Assign items to Active / Inactive / Archive
# ═══════════════════════════════════════════════════════════════

class SystemsAssignView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/systems/assign/
    Body: { item_id, target: 'active'|'inactive', from: 'available'|'active'|'inactive', count }
    Moves `count` units of an item between available/active/inactive buckets.
    """
    def post(self, request, room_slug):
        import json as _json
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        try:
            data = _json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        item_id = data.get('item_id')
        target  = data.get('target')   # 'active', 'inactive', 'available'
        from_   = data.get('from')     # 'available', 'active', 'inactive', 'archive'
        try:
            count = int(data.get('count', 0))
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Count must be at least 1.'}, status=400)

        if target not in ('active', 'inactive', 'available'):
            return JsonResponse({'error': 'Invalid target.'}, status=400)
        if from_ not in ('available', 'active', 'inactive', 'archive'):
            return JsonResponse({'error': 'Invalid source.'}, status=400)
        if count < 1:
            return JsonResponse({'error': 'Count must be at least 1.'}, status=400)

        from django.db import transaction as _tx
        with _tx.atomic():
            item = get_object_or_404(Item.objects.select_for_update(), id=item_id, room=room)

            if from_ == 'available':
                source_count = item.available_count
            elif from_ == 'active':
                source_count = item.active_count
            elif from_ == 'inactive':
                source_count = item.inactive_count
            elif from_ == 'archive':
                source_count = item.archived_count

            if count > source_count:
                return JsonResponse({'error': f'Insufficient count in source {from_}. Available: {source_count}'}, status=400)

            # Deduct from source
            if from_ == 'available':
                item.available_count = max(0, item.available_count - count)
            elif from_ == 'active':
                item.active_count = max(0, item.active_count - count)
            elif from_ == 'inactive':
                item.inactive_count = max(0, item.inactive_count - count)
            elif from_ == 'archive':
                item.archived_count = max(0, item.archived_count - count)
                if item.serviceable_count >= count:
                    item.serviceable_count -= count
                else:
                    item.unserviceable_count = max(0, item.unserviceable_count - (count - item.serviceable_count))
                    item.serviceable_count = 0

            # Add to target
            if target == 'active':
                item.active_count += count
            elif target == 'inactive':
                item.inactive_count += count
            elif target == 'available':
                # Will be recalculated correctly at the end
                pass

            # Recalculate available_count
            item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
            item.in_use = item.active_count
            try:
                item.save(update_fields=['active_count', 'inactive_count', 'available_count', 'in_use', 'archived_count', 'serviceable_count', 'unserviceable_count', 'updated_on'])
            except ValidationError as exc:
                return JsonResponse({'error': exc.messages[0] if hasattr(exc, 'messages') else str(exc)}, status=400)

        return JsonResponse({
            'success': True,
            'message': f'Moved {count} unit(s) of "{item.item_name}" to {target}.',
        })


class SystemsArchiveView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/systems/archive/
    Body: { item_id, from: 'available'|'active'|'inactive', count, category: 'serviceable'|'unserviceable', remark }
    Archives `count` units from the given source bucket.
    """
    def post(self, request, room_slug):
        import json as _json
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        try:
            data = _json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        item_id  = data.get('item_id')
        from_    = data.get('from', 'available')
        try:
            count = int(data.get('count', 0))
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Count must be at least 1.'}, status=400)
        category = data.get('category', 'serviceable')
        remark   = data.get('remark', '').strip()

        if from_ not in ('available', 'active', 'inactive'):
            return JsonResponse({'error': 'Invalid source.'}, status=400)
        if category not in ('serviceable', 'unserviceable'):
            return JsonResponse({'error': 'Invalid category.'}, status=400)
        if count < 1:
            return JsonResponse({'error': 'Count must be at least 1.'}, status=400)

        from django.db import transaction as _tx
        with _tx.atomic():
            item = get_object_or_404(Item.objects.select_for_update(), id=item_id, room=room)

            if from_ == 'available':
                source_count = item.available_count
            elif from_ == 'active':
                source_count = item.active_count
            else:
                source_count = item.inactive_count

            if count > source_count:
                return JsonResponse({'error': 'Insufficient available count'}, status=400)

            # Deduct from source
            if from_ == 'available':
                item.available_count = max(0, item.available_count - count)
            elif from_ == 'active':
                item.active_count = max(0, item.active_count - count)
            else:
                item.inactive_count = max(0, item.inactive_count - count)

            # Add to archived
            item.archived_count += count
            if category == 'serviceable':
                item.serviceable_count += count
            else:
                item.unserviceable_count += count

            # Recalculate available
            item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
            item.in_use = item.active_count
            try:
                item.save(update_fields=[
                    'active_count', 'inactive_count', 'available_count', 'in_use',
                    'archived_count', 'serviceable_count', 'unserviceable_count', 'updated_on'
                ])
            except ValidationError as exc:
                return JsonResponse({'error': exc.messages[0] if hasattr(exc, 'messages') else str(exc)}, status=400)

            # Create Archive record
            Archive.objects.create(
                organisation=item.organisation,
                department=item.department,
                room=room,
                item=item,
                count=count,
                archive_category=category,
                archive_status='archived',
                remark=remark,
            )

        return JsonResponse({
            'success': True,
            'message': f'Archived {count} unit(s) of "{item.item_name}" as {category}.',
        })


class SystemAutomatedCreateView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/systems/create-automated/
    Body: {
        mode: 'general'|'labs',
        system_type: 'single'|'multi',
        system_count: N,
        system_name_format: 'LAB 1 System - 1',
        target_lane: 'active'|'inactive'|'archive',
        source_lane: 'available'|'active'|'inactive',
        items: [{item_id, qty}, ...]
    }
    """
    def post(self, request, room_slug):
        import json as _json
        from django.utils.text import slugify
        from django.db import transaction as _tx
        from django.db.models import F

        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)

        try:
            data = _json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        mode = data.get('mode', 'general')
        system_type = data.get('system_type', 'single')
        system_count = data.get('system_count', 1)
        system_name_format = data.get('system_name_format', '').strip()
        target_lane = data.get('target_lane', 'active')
        source_lane = data.get('source_lane', 'available')
        items_data = data.get('items', [])

        if not system_name_format:
            return JsonResponse({'error': 'System Name / Label is required.'}, status=400)
        if not items_data:
            return JsonResponse({'error': 'No items selected.'}, status=400)

        try:
            system_count = int(system_count)
            if system_count < 1:
                system_count = 1
        except (TypeError, ValueError):
            system_count = 1

        with _tx.atomic():
            # Generate prefix and starting index for labels
            systems_to_create = []
            if mode == 'labs' and system_type == 'multi':
                # e.g., "LAB 1 System - 1" -> prefix="LAB 1 System - ", start_num=1
                parts = system_name_format.rsplit('-', 1)
                if len(parts) == 2 and parts[1].strip().isdigit():
                    prefix = parts[0] + '-'
                    try:
                        start_num = int(parts[1].strip())
                    except ValueError:
                        start_num = 1
                else:
                    prefix = system_name_format + ' - '
                    start_num = 1

                for i in range(system_count):
                    systems_to_create.append(f"{prefix}{start_num + i}".strip())
            else:
                systems_to_create.append(system_name_format)

            # Fetch or create System objects
            created_systems = []
            for sname in systems_to_create:
                sys_obj, _ = System.objects.get_or_create(
                    organisation=room.organisation,
                    room=room,
                    system_name=sname,
                    defaults={'department': room.department}
                )
                created_systems.append(sys_obj)

            # Map target lane to component status
            if target_lane == 'active':
                comp_status = 'active'
            elif target_lane == 'inactive':
                comp_status = 'inactive'
            else:
                comp_status = 'under_maintenance'

            # Allocate items and create components
            for item_info in items_data:
                item_id = item_info.get('item_id')
                qty = int(item_info.get('qty', 0))
                if qty < 1:
                    continue

                item = get_object_or_404(Item.objects.select_for_update(), id=item_id, room=room)

                # Validate source count
                if source_lane == 'available':
                    source_count = item.available_count
                elif source_lane == 'active':
                    source_count = item.active_count
                else:
                    source_count = item.inactive_count

                if qty > source_count:
                    return JsonResponse({
                        'error': f'Insufficient count for "{item.item_name}". Requested: {qty}, Available: {source_count}'
                    }, status=400)

                # Deduct from source bucket
                if source_lane == 'available':
                    item.available_count = max(0, item.available_count - qty)
                elif source_lane == 'active':
                    item.active_count = max(0, item.active_count - qty)
                    item.in_use = item.active_count
                else:
                    item.inactive_count = max(0, item.inactive_count - qty)

                # Add to target bucket
                if target_lane == 'active':
                    item.active_count += qty
                    item.in_use = item.active_count
                elif target_lane == 'inactive':
                    item.inactive_count += qty
                else:
                    # Archive target
                    item.archived_count += qty
                    item.serviceable_count += qty
                    Archive.objects.create(
                        organisation=item.organisation,
                        department=item.department,
                        room=room,
                        item=item,
                        count=qty,
                        archive_category='serviceable',
                        archive_status='archived',
                        remark=f"Automated System Creation ({system_name_format})"
                    )

                item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
                try:
                    item.save()
                except ValidationError as exc:
                    return JsonResponse({'error': exc.messages[0] if hasattr(exc, 'messages') else str(exc)}, status=400)

                # Distribute components across created systems
                num_sys = len(created_systems)
                qty_per_sys = qty // num_sys
                if qty_per_sys < 1 and num_sys > 1:
                    qty_per_sys = 1

                assigned_qty = 0
                for sys_index, sys_obj in enumerate(created_systems):
                    if assigned_qty >= qty:
                        break

                    sys_qty = qty_per_sys
                    if num_sys == 1:
                        sys_qty = qty
                    elif qty_per_sys == 1:
                        sys_qty = 1

                    if assigned_qty + sys_qty > qty:
                        sys_qty = qty - assigned_qty

                    for u in range(sys_qty):
                        # Infer component type from category or item name
                        category_name = item.category.category_name.lower()
                        comp_type = 'other'
                        for c_type, _ in SystemComponent.COMPONENT_TYPES:
                            if c_type in category_name or category_name in c_type:
                                comp_type = c_type
                                break

                        # Unique serial number sequence
                        base_sn = item.serial_number or f"{item.item_name[:10].upper()}-SN"
                        serial_number = f"{base_sn}-{sys_obj.system_name}-{assigned_qty + u + 1}".replace(' ', '')

                        # Create the component directly
                        SystemComponent.objects.create(
                            system=sys_obj,
                            component_item=item,
                            component_type=comp_type,
                            serial_number=serial_number,
                            status=comp_status
                        )

                    assigned_qty += sys_qty

        return JsonResponse({
            'success': True,
            'message': f'Successfully allocated items and configured systems.'
        })


class SystemDeleteView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/systems/<int:pk>/delete/
    Deletes the System, cascading to components.
    """
    def post(self, request, room_slug, pk):
        from django.db import transaction as _tx
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)

        system = get_object_or_404(System, pk=pk, room=room)
        system_name = system.system_name

        with _tx.atomic():
            # Simply delete the system. Cascade deletes related SystemComponents,
            # which fires their post_delete signal in models.py, restoring counts!
            system.delete()

        return JsonResponse({
            'success': True,
            'message': f'System "{system_name}" has been successfully deleted.'
        })


class SystemComponentRevertView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/systems/components/<int:pk>/revert/
    Deletes the specific SystemComponent, returning its count back to loose available stock.
    """
    def post(self, request, room_slug, pk):
        from django.db import transaction as _tx
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)

        component = get_object_or_404(SystemComponent, pk=pk, system__room=room)
        item_name = component.component_item.item_name

        with _tx.atomic():
            component.delete()

        return JsonResponse({
            'success': True,
            'message': f'Component "{item_name}" has been successfully reverted to loose stock.'
        })


class SystemComponentAddView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/systems/<int:pk>/components/add/
    Request JSON: {
        'component_item_id': int,
        'component_type': str,
        'serial_number': str,
        'status': str,
        'quantity': int
    }
    """
    def post(self, request, room_slug, pk):
        import json
        from django.db import transaction as _tx
        from django.db.models import F
        from inventory.models import Room, System, Item, SystemComponent

        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)

        system = get_object_or_404(System, pk=pk, room=room)
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
            
        item_id = data.get('component_item_id')
        comp_type = data.get('component_type')
        serial_number = data.get('serial_number', '').strip()
        status = data.get('status', 'active')
        try:
            quantity = int(data.get('quantity', 1))
        except (ValueError, TypeError):
            quantity = 1
        
        if quantity < 1:
            return JsonResponse({'error': 'Quantity must be at least 1.'}, status=400)
        if not item_id or not comp_type:
            return JsonResponse({'error': 'Item and component type are required.'}, status=400)
            
        with _tx.atomic():
            item = get_object_or_404(Item.objects.select_for_update(), id=item_id, room=room)
            
            if item.available_count < quantity:
                return JsonResponse({'error': f'Insufficient stock. Only {item.available_count} units available for "{item.item_name}".'}, status=400)
                
            base_sn = item.serial_number or f"{item.item_name[:10].upper()}-SN"
            
            # Create components in a loop
            for u in range(quantity):
                if not serial_number:
                    # Generate unique sequential serial number
                    comp_count = SystemComponent.objects.filter(system=system, component_item=item).count()
                    u_serial = f"{base_sn}-{system.system_name}-{comp_count + 1}".replace(' ', '')
                else:
                    if quantity == 1:
                        u_serial = serial_number
                    else:
                        u_serial = f"{serial_number}-{u + 1}".replace(' ', '')
                
                # Verify unique together: system, component_type, serial_number
                if SystemComponent.objects.filter(system=system, component_type=comp_type, serial_number=u_serial).exists():
                    return JsonResponse({'error': f'A component with serial number "{u_serial}" already exists in this system.'}, status=400)
                
                SystemComponent.objects.create(
                    system=system,
                    component_item=item,
                    component_type=comp_type,
                    serial_number=u_serial,
                    status=status
                )
            
            # Dynamically increment matching status counters and decrement available count
            update_fields = {}
            if status == 'active':
                update_fields['active_count'] = F('active_count') + quantity
                update_fields['in_use'] = F('in_use') + quantity
            elif status == 'inactive':
                update_fields['inactive_count'] = F('inactive_count') + quantity
            elif status == 'under_maintenance':
                update_fields['archived_count'] = F('archived_count') + quantity
                update_fields['serviceable_count'] = F('serviceable_count') + quantity
            elif status == 'disposed':
                update_fields['archived_count'] = F('archived_count') + quantity
                update_fields['unserviceable_count'] = F('unserviceable_count') + quantity
                
            update_fields['available_count'] = F('available_count') - quantity
            
            Item.objects.filter(pk=item.pk).update(**update_fields)
            
        return JsonResponse({
            'success': True,
            'message': f'Successfully assigned {quantity} component(s) of "{item.item_name}" to system "{system.system_name}".'
        })


# ═══════════════════════════════════════════════════════════════
# ARCHIVE STATUS UPDATE
# ═══════════════════════════════════════════════════════════════

class ArchiveStatusUpdateView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/archive/<archive_slug>/update-status/
    Body: { status: 'under_maintenance'|'serviced'|'not_serviceable'|'revert' }

    - under_maintenance: stays serviceable, status changes
    - serviced: count returned to available_count, archive record removed
    - not_serviceable: moves count from serviceable to unserviceable on item
    - revert (unserviceable only): count returned to available_count, archive removed
    """
    def post(self, request, room_slug, archive_slug):
        import json as _json
        profile = request.user.profile
        # Allow room incharge, central admin, or sub-admin
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        archive = get_object_or_404(Archive, slug=archive_slug, room=room)
        try:
            logger.info(f"[ArchiveStatusUpdateView] raw body: {request.body}")
            data = _json.loads(request.body)
            logger.info(f"[ArchiveStatusUpdateView] parsed JSON data: {data}")
        except Exception as e:
            logger.error(f"[ArchiveStatusUpdateView] JSON parse failed: {e}")
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        new_status = data.get('status', '').strip()
        try:
            qty = int(data.get('qty', archive.count))
            logger.info(f"[ArchiveStatusUpdateView] parsed qty: {qty}")
        except (ValueError, TypeError) as e:
            logger.warning(f"[ArchiveStatusUpdateView] qty parsing exception: {e}, falling back to archive.count={archive.count}")
            qty = archive.count

        if qty <= 0 or qty > archive.count:
            return JsonResponse({'error': f'Invalid quantity: must be between 1 and {archive.count}.'}, status=400)

        item = archive.item
        from django.db import transaction as _tx

        # Serviceable category options
        if archive.archive_category == 'serviceable':
            if new_status == 'under_maintenance':
                with _tx.atomic():
                    if qty == archive.count:
                        archive.archive_status = 'under_maintenance'
                        archive.save(update_fields=['archive_status', 'updated_on'])
                    else:
                        archive.count = archive.count - qty
                        archive.save(update_fields=['count', 'updated_on'])
                        Archive.objects.create(
                            organisation=archive.organisation,
                            department=archive.department,
                            room=archive.room,
                            item=archive.item,
                            count=qty,
                            archive_type=archive.archive_type,
                            archive_category='serviceable',
                            archive_status='under_maintenance',
                            remark=archive.remark
                        )
                return JsonResponse({'success': True, 'message': f'{qty} unit(s) updated to Under Maintenance.'})

            elif new_status == 'serviced':
                with _tx.atomic():
                    if qty == archive.count:
                        archive.archive_status = 'serviced'
                        archive.save(update_fields=['archive_status', 'updated_on'])
                    else:
                        archive.count = archive.count - qty
                        archive.save(update_fields=['count', 'updated_on'])
                        Archive.objects.create(
                            organisation=archive.organisation,
                            department=archive.department,
                            room=archive.room,
                            item=archive.item,
                            count=qty,
                            archive_type=archive.archive_type,
                            archive_category='serviceable',
                            archive_status='serviced',
                            remark=archive.remark
                        )
                    # Return count to available
                    item.refresh_from_db()
                    item.archived_count = max(0, item.archived_count - qty)
                    item.serviceable_count = max(0, item.serviceable_count - qty)
                    item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
                    item.save(update_fields=['archived_count', 'serviceable_count', 'available_count', 'updated_on'])
                return JsonResponse({'success': True, 'message': f'{qty} unit(s) returned to available stock.'})

            elif new_status == 'not_serviceable':
                with _tx.atomic():
                    if qty == archive.count:
                        archive.archive_category = 'unserviceable'
                        archive.archive_status = 'not_serviceable'
                        archive.save(update_fields=['archive_category', 'archive_status', 'updated_on'])
                    else:
                        archive.count = archive.count - qty
                        archive.save(update_fields=['count', 'updated_on'])
                        Archive.objects.create(
                            organisation=archive.organisation,
                            department=archive.department,
                            room=archive.room,
                            item=archive.item,
                            count=qty,
                            archive_type=archive.archive_type,
                            archive_category='unserviceable',
                            archive_status='not_serviceable',
                            remark=archive.remark
                        )
                    item.refresh_from_db()
                    item.serviceable_count = max(0, item.serviceable_count - qty)
                    item.unserviceable_count += qty
                    item.save(update_fields=['serviceable_count', 'unserviceable_count', 'updated_on'])
                return JsonResponse({'success': True, 'message': f'{qty} unit(s) marked as Not Serviceable.'})

            else:
                return JsonResponse({'error': 'Invalid status for serviceable item.'}, status=400)

        # Unserviceable category — only revert
        elif archive.archive_category == 'unserviceable':
            if new_status == 'revert':
                with _tx.atomic():
                    if qty == archive.count:
                        archive.archive_status = 'serviced'
                        archive.save(update_fields=['archive_status', 'updated_on'])
                    else:
                        archive.count = archive.count - qty
                        archive.save(update_fields=['count', 'updated_on'])
                        Archive.objects.create(
                            organisation=archive.organisation,
                            department=archive.department,
                            room=archive.room,
                            item=archive.item,
                            count=qty,
                            archive_type=archive.archive_type,
                            archive_category='unserviceable',
                            archive_status='serviced',
                            remark=archive.remark
                        )
                    item.refresh_from_db()
                    item.archived_count = max(0, item.archived_count - qty)
                    item.unserviceable_count = max(0, item.unserviceable_count - qty)
                    item.available_count = max(0, item.total_count - item.active_count - item.inactive_count - item.archived_count)
                    item.save(update_fields=['archived_count', 'unserviceable_count', 'available_count', 'updated_on'])
                return JsonResponse({'success': True, 'message': f'{qty} unit(s) reverted to available stock.'})
            else:
                return JsonResponse({'error': 'Only revert is allowed for unserviceable items.'}, status=400)

        return JsonResponse({'error': 'Unknown archive category.'}, status=400)


# ═══════════════════════════════════════════════════════════════
# CONFIGURATIONS — Item-based spec sheets
# ═══════════════════════════════════════════════════════════════

class ConfigurationsListView(LoginRequiredMixin, ListView):
    template_name = 'room_incharge/configurations.html'
    model = ItemConfiguration
    context_object_name = 'configurations'

    def get_queryset(self):
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        return ItemConfiguration.objects.filter(room=room).select_related('item').order_by('-created_on')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        room = get_object_or_404(Room, slug=self.kwargs['room_slug'])
        context['room_slug'] = self.kwargs['room_slug']
        context['room_settings'] = RoomSettings.objects.get_or_create(room=room)[0]
        context['items'] = Item.objects.filter(room=room).order_by('item_name')
        
        # Prefetch systems in this room
        context['systems'] = System.objects.filter(room=room).prefetch_related('systemcomponent_set__component_item').order_by('system_name')
        
        return context


class SaveConfigurationView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/configurations/save/
    Body: { item_ids: [...], configuration_name, configuration_data (JSON string), count }
    Creates one ItemConfiguration per selected item.
    """
    def post(self, request, room_slug):
        import json as _json
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        try:
            data = _json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        item_ids = data.get('item_ids', [])
        cfg_name = data.get('configuration_name', '').strip()
        cfg_data = data.get('configuration_data', '[]')
        count    = int(data.get('count', 1))

        if not item_ids:
            return JsonResponse({'error': 'Select at least one item.'}, status=400)
        if count < 1:
            return JsonResponse({'error': 'Count must be at least 1.'}, status=400)

        # Validate JSON
        try:
            rows = _json.loads(cfg_data)
            if not isinstance(rows, list) or not rows:
                return JsonResponse({'error': 'Add at least one specification row.'}, status=400)
            
            # Validate non-empty specification keys and values
            for row in rows:
                key = str(row.get('spec') or row.get('key', '')).strip()
                val = str(row.get('value', '')).strip()
                if not key or not val:
                    return JsonResponse({'error': 'Specification key and value cannot be empty.'}, status=400)
        except Exception:
            return JsonResponse({'error': 'Invalid specification data.'}, status=400)

        profile = request.user.profile
        created = []
        for item_id in item_ids:
            item = Item.objects.filter(id=item_id, room=room).first()
            if not item:
                continue
            ItemConfiguration.objects.create(
                organisation=room.organisation,
                room=room,
                item=item,
                configuration_name=cfg_name or item.item_name,
                configuration_data=cfg_data,
                count=count,
                created_by=profile,
            )
            created.append(item.item_name)

        if not created:
            return JsonResponse({'error': 'No valid items found.'}, status=400)

        return JsonResponse({'success': True, 'created': created})


class DeleteConfigurationView(LoginRequiredMixin, View):
    """
    POST /rooms/<room_slug>/api/configurations/<slug>/delete/
    """
    def post(self, request, room_slug, cfg_slug):
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        cfg = get_object_or_404(ItemConfiguration, slug=cfg_slug, room=room)
        cfg.delete()
        return JsonResponse({'success': True})


class ItemConfigurationsAPIView(LoginRequiredMixin, View):
    """
    GET /rooms/<room_slug>/api/item-configurations/?item_id=<id>
    Returns all configurations for a given item in this room.
    """
    def get(self, request, room_slug):
        profile = request.user.profile
        if profile.is_central_admin or profile.is_sub_admin:
            room = get_object_or_404(Room, slug=room_slug, organisation=profile.org)
        else:
            room = get_object_or_404(Room, slug=room_slug, incharge=profile)
        item_id = request.GET.get('item_id')
        if not item_id:
            return JsonResponse({'error': 'item_id required'}, status=400)
        cfgs = ItemConfiguration.objects.filter(room=room, item_id=item_id).order_by('-created_on')
        result = [
            {
                'configuration_name': c.configuration_name,
                'configuration_data': c.configuration_data,
                'count': c.count,
                'created_on': c.created_on.strftime('%d %b %Y'),
            }
            for c in cfgs
        ]
        return JsonResponse({'configurations': result})
